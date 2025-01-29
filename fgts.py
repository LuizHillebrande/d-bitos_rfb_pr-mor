from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import pyautogui
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
import os
import re
import pdfplumber
import pandas as pd
import customtkinter as ctk
import json
from time import sleep
import zipfile
import os
import shutil
from thefuzz import process 
import undetected_chromedriver as uc
import openpyxl as opx
from openpyxl import Workbook, load_workbook
from datetime import datetime
from selenium.webdriver.chrome.options import Options


wb_fgts = opx.load_workbook('EMPRESAS FGTS.xlsx')

# Para acessar a planilha
sheet_wb = wb_fgts['Página1']

imagem_alvo = r"certificado_esperado.png"

# Intervalo entre os cliques de tecla
intervalo = 0.5

def salvar_sem_debitos_fgts(razao_social, mes):
    # Criar a pasta se não existir
    pasta_debitos = 'debitos_fgts'
    if not os.path.exists(pasta_debitos):
        os.makedirs(pasta_debitos)

    # Caminho para o arquivo Excel
    arquivo_excel = os.path.join(pasta_debitos, f"sem_debitos_fgts_{mes}.xlsx")
    
    # Se o arquivo já existir, carregar, senão, criar um novo
    if os.path.exists(arquivo_excel):
        wb = load_workbook(arquivo_excel)  # Carregar arquivo existente
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sem Débitos"
        ws.append(["Razão Social", "Mês", "Situação"])  # Adicionar cabeçalho

    # Adicionar nova linha com os dados
    ws.append([razao_social, mes, "Sem débitos FGTS"])
    
    # Salvar o arquivo Excel
    wb.save(arquivo_excel)
    print(f"Empresa {razao_social} salva no Excel com status 'Sem Débitos FGTS' para o mês {mes}.")


def localizar_imagem_na_tela(imagem, confidence=0.8):
    """
    Tenta localizar a imagem na tela.
    :param imagem: Caminho da imagem a ser localizada.
    :param confidence: Nível de confiança para correspondência da imagem.
    :return: Coordenadas da imagem encontrada ou None.
    """
    try:
        # Localiza a imagem na tela
        localizacao = pyautogui.locateOnScreen(imagem, confidence=confidence)
        return localizacao
    except Exception as e:
        print(f"Erro ao localizar imagem: {e}")
        return None

def pressionar_ate_encontrar(imagem, intervalo=0.5):
    """
    Pressiona a seta para baixo até encontrar a imagem na tela.
    :param imagem: Caminho da imagem a ser localizada.
    :param intervalo: Intervalo entre as teclas pressionadas.
    """
    while True:
        localizacao = localizar_imagem_na_tela(imagem)
        if localizacao:
            print(f"Imagem encontrada nas coordenadas: {localizacao}")
            pyautogui.press('enter')
            sleep(3)
            break
        else:
            print("Imagem não encontrada. Pressionando seta para baixo...")
            pyautogui.press('down')
            time.sleep(intervalo)

def pegar_debitos_fgts():
    mes_atual = datetime.now().strftime("%m-%Y")

    # Caminho para o diretório de downloads
    diretorio_download = os.path.join(os.getcwd(), f"debitos_fgts_{mes_atual}")

    # Cria o diretório se não existir
    if not os.path.exists(diretorio_download):
        os.makedirs(diretorio_download)

    # Configurações do Chrome para definir o diretório de download
    chrome_options = Options()
    chrome_options.add_argument(f'--download-default-directory={diretorio_download}')

    # Inicializa o WebDriver com as configurações
    driver = uc.Chrome(options=chrome_options)
    driver.get("https://fgtsdigital.sistema.gov.br/portal/login")
    driver.maximize_window()

    sleep(2)

    entry_gov = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button is-primary entrar']"))
    )
    entry_gov.click()


    entry_certificate = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@id='login-certificate']"))
    )
    entry_certificate.click()
    sleep(3)

    pyautogui.click(738,191, duration=2)
    sleep(3)

    pressionar_ate_encontrar(imagem_alvo, intervalo)

    sleep(5)

    definir = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button is-primary']"))
    )
    definir.click

    for linha in sheet_wb.iter_rows(min_row=2, max_row=500):
        razao_social = linha[1].value
        cnpj = linha[2].value

        try: 
            trocar_perfil = WebDriverWait(driver,10).until(
            EC.element_to_be_clickable((By.XPATH,"//button[@class=' br-button secondary botao-barra-perfil']"))
            )
            trocar_perfil.click
        except:
            print('N tinha trocar perfil')


        dropdown_perfil = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH,"//div[@role='combobox']"))
        )
        dropdown_perfil.click()

        procurador_option = driver.find_element(By.XPATH, "//span[@class='ng-option-label' and text()='Procurador']")

        # Clicando no item
        ActionChains(driver).move_to_element(procurador_option).click().perform()

        input_cnpj = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//input[@class='brx-input medium ng-untouched ng-pristine ng-invalid']"))
        )
        input_cnpj.click()
        input_cnpj.send_keys(cnpj)
        sleep(2)

        selecionar = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button is-primary']"))
        )
        selecionar.click()

        gestao_guias = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//div[contains(@class, 'amplo cardListItem')]//span[contains(text(), 'Gestão de Guias')]"))
        )
        gestao_guias.click()
        sleep(2)

        emissao_guia_rapida = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'amplo cardListItem')]//span[contains(text(), 'Emissão de Guia Rápida')]"))
        )

        # Clicar no elemento
        emissao_guia_rapida.click()
        
        mes = datetime.now().strftime("%m-%Y")

        try:
            # Verificar se o texto contém algo relacionado a "débitos"
            texto_elemento = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'description') and contains(text(), 'débito')]"))
            )
            
            if texto_elemento:
                salvar_sem_debitos_fgts(razao_social, mes)
                print(f"Empresa {razao_social} registrada como sem débitos FGTS para o mês {mes}.")
        except Exception:
            print(f"Empresa {razao_social} possui débitos")
            pesquisar = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button secondary ml-2']"))
            )
            pesquisar.click()

            emitir_guia = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//button[@class='ml-2 br-button primary']"))
            )
            emitir_guia.click()
            print('Emitindo guia para a empresa: ', razao_social)

            try:
                valor_total = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, "//span[contains(text(),',')]"))  # Pegando valores numéricos
                ).text

                vencimento = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, "//span[@title]"))
                ).text

                print(f"Empresa {razao_social} tem débitos no valor de {valor_total} com vencimento para {vencimento}")

                # Criar/abrir arquivo Excel
                nome_arquivo = "debitos_empresas.xlsx"

                try:
                    wb = opx.load_workbook(nome_arquivo)  # Abre o arquivo se existir
                    sheet = wb.active
                except FileNotFoundError:
                    wb = opx.Workbook()  # Cria um novo se não existir
                    sheet = wb.active
                    sheet.append(["Razão Social", "Valor Total", "Vencimento"])  # Cabeçalhos

                # Adiciona os dados
                sheet.append([razao_social, valor_total, vencimento])

                # Salva o arquivo
                wb.save(nome_arquivo)
                print(f"Dados salvos em {nome_arquivo}")

            except Exception as e:
                print(f"Erro ao extrair os dados: {e}")

        driver.get('https://fgtsdigital.sistema.gov.br/portal/servicos')


        sleep(2)
        
    driver.quit()
pegar_debitos_fgts()
