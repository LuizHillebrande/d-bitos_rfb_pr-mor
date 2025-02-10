from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import pyautogui
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
import os
import re
import pdfplumber
import pandas as pd
import customtkinter as ctk
import json
from time import sleep
import zipfile
import os
import shutil
from thefuzz import process 
import undetected_chromedriver as uc
import openpyxl as opx
from openpyxl import Workbook, load_workbook
from datetime import datetime
from selenium.webdriver.chrome.options import Options

wb_fgts = opx.load_workbook('EMPRESAS FGTS.xlsx')

# Para acessar a planilha
sheet_wb = wb_fgts['Página1']

imagem_alvo = r"certificado_esperado.png"

# Intervalo entre os cliques de tecla
intervalo = 0.5

def salvar_sem_debitos_fgts(razao_social, mes):
    # Criar a pasta se não existir
    pasta_debitos = 'debitos_fgts'
    if not os.path.exists(pasta_debitos):
        os.makedirs(pasta_debitos)

    # Caminho para o arquivo Excel
    arquivo_excel = os.path.join(pasta_debitos, f"sem_debitos_fgts_{mes}.xlsx")
    
    # Se o arquivo já existir, carregar, senão, criar um novo
    if os.path.exists(arquivo_excel):
        wb = load_workbook(arquivo_excel)  # Carregar arquivo existente
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sem Débitos"
        ws.append(["Razão Social", "Mês", "Situação"])  # Adicionar cabeçalho

    # Adicionar nova linha com os dados
    ws.append([razao_social, mes, "Sem débitos FGTS"])
    
    # Salvar o arquivo Excel
    wb.save(arquivo_excel)
    print(f"Empresa {razao_social} salva no Excel com status 'Sem Débitos FGTS' para o mês {mes}.")


def localizar_imagem_na_tela(imagem, confidence=0.8):
    """
    Tenta localizar a imagem na tela.
    :param imagem: Caminho da imagem a ser localizada.
    :param confidence: Nível de confiança para correspondência da imagem.
    :return: Coordenadas da imagem encontrada ou None.
    """
    try:
        # Localiza a imagem na tela
        localizacao = pyautogui.locateOnScreen(imagem, confidence=confidence)
        
        if localizacao:
            # Move o mouse para as coordenadas do centro da imagem
            pyautogui.moveTo(pyautogui.center(localizacao))
            sleep(1)
            pyautogui.click()
            return localizacao
        else:
            print("Imagem não encontrada.")
            return None
    except Exception as e:
        print(f"Erro ao localizar imagem: {e}")
        return None


def pressionar_ate_encontrar(imagem, intervalo=0.5):
    """
    Pressiona a seta para baixo até encontrar a imagem na tela.
    :param imagem: Caminho da imagem a ser localizada.
    :param intervalo: Intervalo entre as teclas pressionadas.
    """
    while True:
        localizacao = localizar_imagem_na_tela(imagem)
        if localizacao:
            print(f"Imagem encontrada nas coordenadas: {localizacao}")
            pyautogui.press('enter')
            sleep(3)
            break
        else:
            print("Imagem não encontrada. Pressionando seta para baixo...")
            pyautogui.press('down')
            time.sleep(intervalo)

def extrair_cnpj_nome_empresa(driver):
    try:
        # Localiza o elemento que contém o CNPJ e nome da empresa
        dados_perfil = driver.find_element(By.XPATH, "//span[@class='dados-perfil']")
        
        # Extrai o texto
        texto = dados_perfil.text.strip()
        
        # Remove o prefixo "Empregador: " caso exista
        if texto.startswith("Empregador: "):
            texto = texto.replace("Empregador: ", "")
        
        # Verifica se o formato está correto (deve conter ' | ')
        if ' | ' in texto:
            cnpj, nome_empresa = texto.split(' | ')
            
            # Limpa o CNPJ (remove pontuação, se necessário)
            cnpj_limpo = cnpj.replace('.', '').replace('-', '').replace('/', '')
            
            # Formata o nome da empresa
            nome_empresa_formatado = nome_empresa.strip()
            
            # Cria a string no formato desejado
            resultado = f"{cnpj_limpo}_{nome_empresa_formatado}"
            
            return resultado
        else:
            print(f"Formato inesperado do texto extraído: {texto}")
            return None
    
    except Exception as e:
        print(f"Erro ao extrair CNPJ e nome da empresa: {e}")
        return None



def pegar_debitos_fgts():
    mes_atual = datetime.now().strftime("%m-%Y")

    # Caminho para o diretório de downloads
    diretorio_download = os.path.join(os.getcwd(), f"debitos_fgts_{mes_atual}")

    # Cria o diretório se não existir
    if not os.path.exists(diretorio_download):
        os.makedirs(diretorio_download)

    # Configurações do Chrome para definir o diretório de download
    chrome_options = Options()
    chrome_options.add_argument(f'--download-default-directory={diretorio_download}')

    # Inicializa o WebDriver com as configurações
    driver = uc.Chrome(options=chrome_options)
    driver.get("https://fgtsdigital.sistema.gov.br/portal/login")
    driver.get("https://fgtsdigital.sistema.gov.br/portal/login")
    driver.maximize_window()

    sleep(2)

    entry_gov = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button is-primary entrar']"))
    )
    entry_gov.click()


    entry_certificate = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@id='login-certificate']"))
    )
    entry_certificate.click()
    sleep(5)

    try:
        hcaptcha_iframe = driver.find_element(By.XPATH, "//iframe[contains(@src, 'hcaptcha')]")
        if hcaptcha_iframe:
            print("hCaptcha detectado via Selenium.")
            driver.quit()
            sleep(10)
            pegar_debitos_fgts()
            return
    except Exception as e:
        # Se não encontrar, continua
        pass


    pyautogui.click(738,191, duration=2)
    sleep(5)

    pressionar_ate_encontrar(imagem_alvo, intervalo)

    sleep(3)

    try:
        aceitar_cookies = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button primary small']"))
        )
        aceitar_cookies.click()
    except Exception as e:
        print('Erro ao aceitar cookies', e)

    definir = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button is-primary']"))
    )
    definir.click()
    print('Clicado em definir')


    for linha in sheet_wb.iter_rows(min_row=2, max_row=500):
        cnpj = linha[2].value

        from selenium.common.exceptions import StaleElementReferenceException

        try:
            for _ in range(3):  # Tenta encontrar o botão até 3 vezes
                try:
                    trocar_perfil = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'botao-barra-perfil')]"))
                    )
                    trocar_perfil.click()
                    print('Clicado em "Trocar Perfil"')
                    break  # Sai do loop se conseguiu clicar
                except StaleElementReferenceException:
                    print("Elemento ficou obsoleto, tentando novamente...")
        except Exception as e:
            print(f"Erro ao clicar em 'Trocar Perfil': {e}")


        dropdown_perfil = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH,"//div[@role='combobox']"))
        )
        dropdown_perfil.click()

        procurador_option = driver.find_element(By.XPATH, "//span[@class='ng-option-label' and text()='Procurador']")

        # Clicando no item
        ActionChains(driver).move_to_element(procurador_option).click().perform()

        input_cnpj = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//input[@class='brx-input medium ng-untouched ng-pristine ng-invalid']"))
        )
        input_cnpj.click()
        input_cnpj.send_keys(cnpj)
        sleep(2)

        selecionar = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button is-primary']"))
        )
        selecionar.click()
        sleep(1)

        resultado_cnpj_nome = extrair_cnpj_nome_empresa(driver)
    
        if resultado_cnpj_nome:
            print(f"Resultado extraído: {resultado_cnpj_nome}")
        else:
            print("Não foi possível extrair o CNPJ e nome da empresa.")
        
        razao_social = resultado_cnpj_nome

        consultas_empregador = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//div[contains(@class, 'amplo cardListItem')]//span[contains(text(), 'Consultas do Empregador')]"))
        )
        consultas_empregador.click()
        sleep(2)

        consultar_competencia = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'amplo cardListItem')]//span[contains(text(), 'Consulta de Competências de Referência')]"))
        )

        # Clicar no elemento
        consultar_competencia.click()

        #elemento completo
        elemento = driver.find_element(By.XPATH, "//label[@for='completo']")
        actions = ActionChains(driver)
        actions.move_to_element(elemento).click().perform()
        sleep(1)

        #tirar o regular
        regular = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//label[@for='regular']"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(regular).click().perform()
        sleep(3)

        filtrar = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button secondary']"))
        )
        filtrar.click()
        

        sleep(3)

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        sleep(2)  # Aguarde um pouco para garantir que o conteúdo foi carregado

        # Tentar localizar o elemento
        #elemento = WebDriverWait(driver, 10).until(
            #EC.element_to_be_clickable((By.CSS_SELECTOR, ".ng-arrow-wrapper"))
        #)
        elemento = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH, "(//div[@class='ng-input'])[1]"))
        )

        # Realizar a ação no elemento
        ActionChains(driver).move_to_element(elemento).click().perform()
        sleep(2)

        #checkbox_10 = WebDriverWait(driver, 10).until(
            #EC.element_to_be_clickable((By.XPATH, "(//div[@class='ng-value']//span[@class='ng-value-label' and text()='10'])[1]"))
        #)
        #checkbox_10.click()  # Clicando no valor "10"
        
        campo_texto = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//input[@aria-autocomplete='list']"))
        )

        # Clique no campo para abrir a lista
        campo_texto.click()

        # Digite o valor que você deseja selecionar (por exemplo, '100')
        campo_texto.send_keys("100")
        pyautogui.press('enter')
        sleep(3)

        dados = []

        # Encontra todas as linhas da tabela
        linhas = driver.find_elements(By.XPATH, "//datatable-body-row")

        for linha in linhas:
            try:
                # Extrai o mês de referência (supondo que seja a terceira célula)
                mes_ref = linha.find_element(By.XPATH, ".//datatable-body-cell[3]").text.strip()
                
                # Extrai o valor do débito (localizando pelo estilo de cor vermelha)
                valor_debito = linha.find_element(By.XPATH, ".//span[@style='color: #b30000; font-weight: 600;']").text.strip()

                # Remove caracteres indesejados no valor
                valor_debito = valor_debito.replace(".", "").replace(",", ".")

                # Adiciona os dados à lista
                dados.append([razao_social, mes_ref, float(valor_debito)])

            except Exception as e:
                print(f'Empresa {razao_social} sem débitos')
                print(f"Erro ao processar linha: {e}")

        # Criando um DataFrame com os novos dados
        df_novos_dados = pd.DataFrame(dados, columns=["Nome da Empresa", "Mês Ref.", "Valor Débitos"])

        # Caminho do arquivo Excel
        arquivo_excel = "debitos_fgts.xlsx"

        # Verificar se o arquivo já existe
        if os.path.exists(arquivo_excel):
            # Se o arquivo já existe, carregar os dados existentes
            df_existente = pd.read_excel(arquivo_excel)

            # Concatenar os dados novos com os dados existentes
            df_total = pd.concat([df_existente, df_novos_dados], ignore_index=True)

            # Salvar o DataFrame atualizado
            df_total.to_excel(arquivo_excel, index=False)
        else:
            # Se o arquivo não existe, criar um novo arquivo com os dados
            df_novos_dados.to_excel(arquivo_excel, index=False)

        print("Arquivo Excel atualizado com sucesso!")

        mes = datetime.now().strftime("%m-%Y")

        driver.get('https://fgtsdigital.sistema.gov.br/portal/servicos')


        sleep(2)
        
    driver.quit()
pegar_debitos_fgts()

'''

def criar_interface_fgts():
    # Limpa somente o conteúdo do info_frame (área dinâmica)
    for widget in info_frame.winfo_children():
        widget.destroy()

    # Cria um frame para a interface do FGTS Digital
    fgts_frame = ctk.CTkFrame(master=info_frame, width=400, height=500, corner_radius=20, fg_color="#2E2E2E")
    fgts_frame.place(relx=0.5, rely=0.5, anchor="center")

    # Título para a interface do FGTS Digital
    titulo_fgts = ctk.CTkLabel(
        master=fgts_frame,
        text="FGTS Digital",
        font=("Arial", 24, "bold"),
        text_color="#00A3FF"
    )
    titulo_fgts.pack(pady=20)

    # Instrução ou informações adicionais
    instrucao_fgts = ctk.CTkLabel(
        master=fgts_frame,
        text="Clique no botão abaixo para extrair débitos FGTS.",
        font=("Arial", 16),
        text_color="white"
    )
    instrucao_fgts.pack(pady=10)

    # Botão para chamar a função de extração dos débitos FGTS
    extrair_fgts_button = ctk.CTkButton(
        master=fgts_frame,
        text="Extrair Débitos FGTS",
        command=pegar_debitos_fgts,  # Função que você já definiu para extrair os débitos
        width=300,
        height=50,
        corner_radius=10,
        fg_color="#00A3FF",
        text_color="white",
        hover_color="#0088CC"
    )
    extrair_fgts_button.pack(pady=20)
'''