import customtkinter as ctk
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tkinter import messagebox
from PIL import Image
import pyautogui
from time import sleep
import pyperclip
import openpyxl
import os
import re
import pandas as pd
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from Levenshtein import distance

# Credenciais padrão
DEFAULT_EMAIL = "legal@contabilprimor.com.br"
DEFAULT_SENHA = "q7ne5k0la0VJ"


def extrair_nomes_empresas():
    pasta_debitos = "debitos"
    pasta_saida = "nomes_empresas"
    arquivo_saida = os.path.join(pasta_saida, "empresas.xlsx")

    # Garante que a pasta de saída existe
    os.makedirs(pasta_saida, exist_ok=True)

    dados = []

    # Iterar sobre os arquivos na pasta 'debitos'
    for arquivo in os.listdir(pasta_debitos):
        if arquivo.endswith(".pdf"):
            partes = arquivo.split("_")
            if len(partes) >= 2:
                cnpj = partes[0]
                nome_empresa = " ".join(partes[1:]).replace(".pdf", "").strip()
                dados.append([nome_empresa, cnpj])

    # Verifica se encontrou algum dado
    if not dados:
        print("Nenhum dado encontrado.")
        return

    # Criar um DataFrame e salvar como Excel
    df = pd.DataFrame(dados, columns=["Nome Empresa", "CNPJ"])
    df.to_excel(arquivo_saida, index=False)

    print(f"Arquivo salvo em: {arquivo_saida}")



# Função para iniciar o WebDriver com as credenciais
def iniciar_webdriver(email, senha):
    empresas_nao_enviadas = []
    total_empresas = contar_pdfs()  # Agora conta os arquivos PDF na pasta 'debitos'
    total_enviadas = 0
    #relatório de pendências fiscais
    #extrair_nomes_empresas()
    excel_msg = openpyxl.load_workbook("nomes_empresas\empresas.xlsx")
    sheet_excel_msg = excel_msg.active
    try:
        driver = webdriver.Chrome()
        driver.get("https://app.digiliza.com.br/login")
        driver.maximize_window()


        for linha in sheet_excel_msg.iter_rows(min_row=2, max_row=100):
            contador_vazios = 0
            if linha!=2:
                try:
                    driver.quit()  # Fecha antes de iniciar (se já estiver rodando)
                    # Inicia um novo WebDriver
                    driver = webdriver.Chrome()  
                    driver.get("https://app.digiliza.com.br/login")
                    driver.maximize_window()
                    input_email = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@id='email']"))
                    )
                    input_email.send_keys(email)

                    input_senha = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@id='password']"))
                    )
                    input_senha.send_keys(senha)

                    botao_login = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//button[@type='submit']"))
                    )
                    botao_login.click()
                except:
                    pass
            
    
            data_vcto_input_padrao = datetime.now().strftime("27/%m/%Y")
            print('Lopacarai',data_vcto_input_padrao)
            data_atual = datetime.now().strftime("%d/%m/%y")
            competencia_padrao = datetime.now().strftime("%m/%y")
            nome_empresa = linha[0].value
            cnpj = linha[1].value
            input_complemento = f'Análise de pendências da competência {data_atual}'

            if not nome_empresa or not cnpj:
                driver.quit()
                messagebox.showinfo('Sucesso!', 'Mensagens enviadas com sucesso!')
                break
            # Esperar até que o botão esteja presente
            incluir_tarefa_avulsa = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[@href='#modalTarefaAvulsa']"))
            )

            # Usar JavaScript para garantir o clique se necessário
            driver.execute_script("arguments[0].click();", incluir_tarefa_avulsa)

            sleep(3)

            select_model = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Selecione um Modelo']"))
            )
            select_model.click()
            sleep(2)

            texto = "relatório de pendências fiscais"
            pyperclip.copy(texto)
            pyautogui.hotkey("ctrl", "v")
            sleep(2)
            pyautogui.press('enter')

            cliente = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Selecione um Cliente']"))
            )
            cliente.click()
            cliente.send_keys(nome_empresa)
            sleep(2)
            pyautogui.press('enter')

            complemento = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@id='modalAddTaskComplemento']"))
            )
            complemento.click()
            complemento.send_keys(input_complemento)

            data_vcto_digiliza = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@type='date'][1]"))
            )
            data_vcto_digiliza.click()
            data_vcto_digiliza.clear()
            sleep(1)
            data_vcto_digiliza.send_keys(data_vcto_input_padrao)

            competencia = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='__/____']"))
            )
            competencia.click()
            competencia.clear()
            competencia.send_keys(competencia_padrao)
            
            '''
            #n precisa disso pq sempre vai ser primor contabil o responsavel
            responsavel = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//div[@class='v-select vs--single vs--searchable mh-sm scrollbar scrollbar-3']"))
            )
            responsavel.click()
            responsavel.clear()
            texto_responsavel = 'Prímor Contábil'
            pyperclip.copy(texto_responsavel)
            pyautogui.hotkey("ctrl", "v")
            print('dei ctrl v')
            sleep(15)
            pyautogui.press('enter')
            '''

            sleep(5)

            botao_salvar = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Salvar e ir à(s) Tarefa(s)')]"))
            )
            botao_salvar.click()
            sleep(3)

            # Supondo que 'driver' já esteja inicializado
            try:
                # Espera até 2 segundos para encontrar o título do erro
                WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.XPATH, "//h2[@id='swal2-title' and contains(text(), 'Erro ao criar tarefa(s)')]"))
                )
                
                # Se encontrar, espera o botão "OK" e clica nele
                btn_ok = WebDriverWait(driver, 2).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'swal2-confirm')]"))
                )
                btn_ok.click()
                print("Botão 'OK' clicado com sucesso.")

                empresas_nao_enviadas.append(f"{nome_empresa} ({cnpj})")
                continue
            except Exception as e:  
                if "WinError 6" in str(e) or "chrome not reachable" in str(e):
                    print("ChromeDriver fechou inesperadamente. Reiniciando...")
                    driver.quit()
                    driver = webdriver.Chrome()  # Inicia um novo WebDriver
                    print("Mensagem de erro não encontrada dentro do tempo limite.")


            #Acessando o elemento com nome da empresa incrementado c f''
            

            # Captura todas as empresas da tabela
            nomes_encontrados = [e.text for e in driver.find_elements(By.XPATH, "//td")]
            print('Nomes encontrados\n')
            # Encontra o nome mais próximo
            nome_correto = min(nomes_encontrados, key=lambda x: distance(x.lower(), nome_empresa.lower()))
            print('Nome correto', nome_correto)

            # Clica no nome encontrado
            elemento_empresa = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, f"//td[contains(text(), '{nome_correto}')]"))
            )
            elemento_empresa.click()

            sleep(3)

            botao_documentos_svg = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-transparent rounded-0 text-primary flex-fill']"))
            )
            botao_documentos_svg.click()
            print('Acessei a aba de documentos')
            sleep(3)

            # Diretório onde estão os PDFs
            pasta_debitos = os.path.join(os.getcwd(), 'debitos')

            # Construir padrão do nome do arquivo
            padrao_arquivo = f"{cnpj}_*.pdf"

            # Procurar o arquivo correto na pasta
            arquivos_encontrados = [f for f in os.listdir(pasta_debitos) if f.startswith(f"{cnpj}_") and f.endswith(".pdf")]

            if arquivos_encontrados:
                caminho_arquivo = os.path.join(pasta_debitos, arquivos_encontrados[0])  # Pega o primeiro encontrado


                # Localizar o campo de upload oculto e enviar o arquivo
                campo_upload = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@type='file']"))
                )
                campo_upload.send_keys(caminho_arquivo)

                print(f"Arquivo {caminho_arquivo} anexado com sucesso!")

                try:
                    aceitar_pop_up = WebDriverWait(driver,3).until(
                        EC.element_to_be_clickable((By.XPATH,"//button[@class='swal2-confirm px-3 py-2 swal2-styled swal2-default-outline']"))
                    )
                    aceitar_pop_up.click()

                except:
                    print('n tinha pop up')
                sleep(2)
            else:
                print(f"Nenhum arquivo encontrado para {nome_empresa} ({cnpj})")

            # Esperar o upload ser processado (caso tenha carregamento)
            sleep(5) 

            enviar_msg = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-sm btn-transparent px-1 py-0'] [1]"))
            )
            enviar_msg.click()

            click_edit_mensagem = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//div[@class='ql-editor ql-blank']"))
            )
            click_edit_mensagem.click()

            # Abrir o arquivo mensagens.xlsx
            excel_mensagens = openpyxl.load_workbook("mensagens.xlsx")
            sheet_mensagens = excel_mensagens.active

            # Procurar o CNPJ na coluna 0 e pegar a mensagem da coluna 1
            mensagem_personalizada = None
            for linha in sheet_mensagens.iter_rows(min_row=2, max_row=10):
                cnpj_planilha = str(linha[0].value)
                print(f'Cnpj {cnpj_planilha}')
                mensagem = str(linha[1].value)
                #print(f'Mensagem {mensagem}')
                if str(cnpj).strip() == str(cnpj_planilha).strip():  # Comparar sem espaços extras
                    mensagem_personalizada = mensagem
                    break

            # Se encontrou a mensagem, colar no campo de edição
            if mensagem_personalizada:
                click_edit_mensagem = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, "//div[@class='ql-editor ql-blank']"))
                )
                click_edit_mensagem.click()

                # Copiar e colar usando pyperclip + pyautogui
                click_edit_mensagem.send_keys(mensagem_personalizada)
                print("Colei")
                sleep(2)

                botoes = driver.find_elements(By.XPATH, "//button[text()='Salvar']")
                for botao in botoes:
                    if botao.is_displayed():
                        botao.click()
                        break

                print('Cliquei em salvar')
                sleep(5)
                
                voltar = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-transparent rounded-0 text-secondary flex-fill']"))
                )
                voltar.click()

                iniciar_etapa = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-sm btn-transparent btn-hover-semi-transparent text-white']"))
                )
                iniciar_etapa.click()
                sleep(5)

                concluir_etapa = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-sm btn-transparent btn-hover-semi-transparent text-white big-chungus']"))
                )
                concluir_etapa.click()
                sleep(3)

                botao_ok = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='swal2-confirm px-3 py-2 swal2-styled']"))
                )
                botao_ok.click()

                total_enviadas += 1

                '''
                concluir_e_enviar = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[@class='btn btn-success' and contains(text(), 'Concluir e enviar')]"))
                )
                concluir_e_enviar.click()

                botao_ok = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='swal2-confirm px-3 py-2 swal2-styled']"))
                )
                botao_ok.click()
                '''
                botoes_fechar = driver.find_elements(By.XPATH, "//button[@data-bs-dismiss='modal']")
                for botao in botoes_fechar:
                    
                    if botao.is_displayed():  # Verifica se o botão está visível na tela
                        botao.click()
                        break  # Para após clicar no primeiro botão visível

                

                sleep(5)

                print(f"Mensagem enviada para {nome_empresa}: {mensagem_personalizada}")
            else:
                empresas_nao_enviadas.append(f"{nome_empresa} ({cnpj})")
                '''
                botoes_fechar = driver.find_elements(By.XPATH, "//button[@data-bs-dismiss='offcanvas']")
                for botao in botoes_fechar:
                    print('Nao achei a mensagem!')
                    if botao.is_displayed():  # Verifica se o botão está visível na tela
                        botao.click()
                        break  # Para após clicar no primeiro botão visível
                '''
                print(f"⚠ Nenhuma mensagem personalizada encontrada para {nome_empresa} ({cnpj})")

        if empresas_nao_enviadas:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Nome da Empresa", "CNPJ"])
            for empresa in empresas_nao_enviadas:
                nome, cnpj = empresa.split(" (")
                cnpj = cnpj[:-1]  # Removendo o parêntese final
                ws.append([nome, cnpj])
    
            wb.save("empresas_nao_enviadas.xlsx")
            print("Arquivo 'empresas_nao_enviadas.xlsx' salvo com sucesso.")

        # Enviar email com os resultados
        enviar_email(empresas_nao_enviadas, total_empresas, total_enviadas)
        messagebox.showinfo("Sucesso", "Login realizado com sucesso!")
        sleep(50)

        if driver.service.is_connectable():
            driver.quit()
    except Exception as e:
        print("Erro", f"Falha no login: {e}")

def contar_pdfs():
    pasta = "debitos"
    return len([f for f in os.listdir(pasta) if f.endswith(".pdf")])

def enviar_email(empresas_nao_enviadas, total_empresas, total_enviadas):
    remetente_email = "luizhill.dev@gmail.com"  # Altere para seu email
    senha_email = "nqlf fgch thrs kpht"  # Use app password se for Gmail
    destinatarios = ["luiz.logika@gmail.com", "luiz.hillebrande1505@gmail.com"]

    saudacao = "Bom dia" if datetime.now().hour < 12 else "Boa tarde"
    lista_empresas = "\n".join(empresas_nao_enviadas)

    mensagem = f"""{saudacao},  
    Informo que as seguintes empresas não foram enviadas:  
    {lista_empresas}  

    Havia um total de {total_empresas} empresas (arquivos PDF na pasta 'debitos'), foram enviadas {total_enviadas}.
    """

    msg = MIMEMultipart()
    msg["From"] = remetente_email
    msg["To"] = ", ".join(destinatarios)
    msg["Subject"] = "Empresas não enviadas"

    msg.attach(MIMEText(mensagem, "plain"))

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)  # Alterar se não for Gmail
        server.starttls()
        server.login(remetente_email, senha_email)
        server.sendmail(remetente_email, destinatarios, msg.as_string())
        server.quit()
        print("Email enviado com sucesso!")
    except Exception as e:
        print(f"Erro ao enviar email: {e}")

iniciar_webdriver(email='legal@contabilprimor.com.br',senha='q7ne5k0la0VJ')
# Criando a interface gráfica
'''
def criar_interface():
    def fazer_login():
        email = email_entry.get()
        senha = senha_entry.get()
        iniciar_webdriver(email, senha)

    def toggle_password():
        """ Alterna entre exibir ou ocultar a senha """
        if senha_entry.cget("show") == "*":
            senha_entry.configure(show="")
            toggle_button.configure(image=eye_open)
        else:
            senha_entry.configure(show="*")
            toggle_button.configure(image=eye_closed)

    app = ctk.CTk()
    app.title("Login - Digiliza")
    app.geometry(f"{app.winfo_screenwidth()}x{app.winfo_screenheight()}+0+0")  # Tela cheia
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    # Fundo estilizado
    bg_frame = ctk.CTkFrame(master=app, fg_color="#1E1E1E")
    bg_frame.pack(fill="both", expand=True)

    # Container do login centralizado
    frame = ctk.CTkFrame(master=bg_frame, width=400, height=500, corner_radius=20, fg_color="#2E2E2E")
    frame.place(relx=0.5, rely=0.5, anchor="center")

    titulo = ctk.CTkLabel(master=frame, text="Login no Digiliza", font=("Arial", 24, "bold"), text_color="#00A3FF")
    titulo.pack(pady=20)

    email_label = ctk.CTkLabel(master=frame, text="E-mail:", text_color="white")
    email_label.pack()
    email_entry = ctk.CTkEntry(master=frame, width=300, height=40, corner_radius=10)
    email_entry.insert(0, DEFAULT_EMAIL)
    email_entry.pack(pady=5)

    senha_label = ctk.CTkLabel(master=frame, text="Senha:", text_color="white")
    senha_label.pack()

    # Campo de senha com botão de exibição
    senha_frame = ctk.CTkFrame(master=frame, fg_color="transparent")
    senha_frame.pack()

    senha_entry = ctk.CTkEntry(master=senha_frame, width=260, height=40, corner_radius=10, show="*")
    senha_entry.insert(0, DEFAULT_SENHA)
    senha_entry.pack(side="left", pady=5)

    # Ícones para alternar a visibilidade da senha
    eye_open = ctk.CTkImage(light_image=Image.open("imgs\eye_open.png"), size=(24, 24))
    eye_closed = ctk.CTkImage(light_image=Image.open("imgs\eye_closed.png"), size=(24, 24))

    toggle_button = ctk.CTkButton(master=senha_frame, width=40, height=40, text="", image=eye_closed,
                                  fg_color="transparent", hover_color="#444", command=toggle_password)
    toggle_button.pack(side="right", padx=5)

    # Botão estilizado
    def on_enter(e):
        login_button.configure(fg_color="#0088CC")  # Azul mais vibrante ao passar o mouse

    def on_leave(e):
        login_button.configure(fg_color="#00A3FF")  # Retorna ao azul original

    login_button = ctk.CTkButton(master=frame, text="Login", command=fazer_login, 
                                 width=300, height=50, corner_radius=10, fg_color="#00A3FF", text_color="white",
                                 hover_color="#0088CC")
    login_button.pack(pady=20)

    login_button.bind("<Enter>", on_enter)
    login_button.bind("<Leave>", on_leave)

    app.mainloop()

# Iniciar a interface
criar_interface()
'''
