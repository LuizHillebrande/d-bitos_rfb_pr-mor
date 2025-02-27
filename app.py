from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import pyautogui
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
import os
import re
import pdfplumber
import pandas as pd
import customtkinter as ctk
import json
from time import sleep
import zipfile
import os
import shutil
from thefuzz import process 
from selenium.webdriver.common.by import By

import os
import pandas as pd
import re
from fuzzywuzzy import process
from datetime import datetime
import textwrap

#FGTS DIGITAL
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import pyautogui
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
import os
import re
import pdfplumber
import pandas as pd
import customtkinter as ctk
import json
from time import sleep
import zipfile
import os
import shutil
from thefuzz import process 
import undetected_chromedriver as uc
import openpyxl as opx
from openpyxl import Workbook, load_workbook
from datetime import datetime
from selenium.webdriver.chrome.options import Options

wb_fgts = opx.load_workbook('EMPRESAS FGTS.xlsx')

# Para acessar a planilha
sheet_wb = wb_fgts['Página1']

imagem_alvo = r"certificado_esperado.png"

# Intervalo entre os cliques de tecla
intervalo = 0.5

def salvar_sem_debitos_fgts(razao_social, mes):
    # Criar a pasta se não existir
    pasta_debitos = 'debitos_fgts'
    if not os.path.exists(pasta_debitos):
        os.makedirs(pasta_debitos)

    # Caminho para o arquivo Excel
    arquivo_excel = os.path.join(pasta_debitos, f"sem_debitos_fgts_{mes}.xlsx")
    
    # Se o arquivo já existir, carregar, senão, criar um novo
    if os.path.exists(arquivo_excel):
        wb = load_workbook(arquivo_excel)  # Carregar arquivo existente
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sem Débitos"
        ws.append(["Razão Social", "Mês", "Situação"])  # Adicionar cabeçalho

    # Adicionar nova linha com os dados
    ws.append([razao_social, mes, "Sem débitos FGTS"])
    
    # Salvar o arquivo Excel
    wb.save(arquivo_excel)
    print(f"Empresa {razao_social} salva no Excel com status 'Sem Débitos FGTS' para o mês {mes}.")


def localizar_imagem_na_tela(imagem, confidence=0.8):
    """
    Tenta localizar a imagem na tela.
    :param imagem: Caminho da imagem a ser localizada.
    :param confidence: Nível de confiança para correspondência da imagem.
    :return: Coordenadas da imagem encontrada ou None.
    """
    try:
        # Localiza a imagem na tela
        localizacao = pyautogui.locateOnScreen(imagem, confidence=confidence)
        
        if localizacao:
            # Move o mouse para as coordenadas do centro da imagem
            pyautogui.moveTo(pyautogui.center(localizacao))
            sleep(1)
            pyautogui.click()
            return localizacao
        else:
            print("Imagem não encontrada.")
            return None
    except Exception as e:
        print(f"Erro ao localizar imagem: {e}")
        return None


def pressionar_ate_encontrar(imagem, intervalo=0.5):
    """
    Pressiona a seta para baixo até encontrar a imagem na tela.
    :param imagem: Caminho da imagem a ser localizada.
    :param intervalo: Intervalo entre as teclas pressionadas.
    """
    while True:
        localizacao = localizar_imagem_na_tela(imagem)
        if localizacao:
            print(f"Imagem encontrada nas coordenadas: {localizacao}")
            pyautogui.press('enter')
            sleep(3)
            break
        else:
            print("Imagem não encontrada. Pressionando seta para baixo...")
            pyautogui.press('down')
            time.sleep(intervalo)

def extrair_cnpj_nome_empresa(driver):
    try:
        # Localiza o elemento que contém o CNPJ e nome da empresa
        dados_perfil = driver.find_element(By.XPATH, "//span[@class='dados-perfil']")
        
        # Extrai o texto
        texto = dados_perfil.text.strip()
        
        # Remove o prefixo "Empregador: " caso exista
        if texto.startswith("Empregador: "):
            texto = texto.replace("Empregador: ", "")
        
        # Verifica se o formato está correto (deve conter ' | ')
        if ' | ' in texto:
            cnpj, nome_empresa = texto.split(' | ')
            
            # Limpa o CNPJ (remove pontuação, se necessário)
            cnpj_limpo = cnpj.replace('.', '').replace('-', '').replace('/', '')
            
            # Formata o nome da empresa
            nome_empresa_formatado = nome_empresa.strip()
            
            # Cria a string no formato desejado
            resultado = f"{cnpj_limpo}_{nome_empresa_formatado}"
            
            return resultado
        else:
            print(f"Formato inesperado do texto extraído: {texto}")
            return None
    
    except Exception as e:
        print(f"Erro ao extrair CNPJ e nome da empresa: {e}")
        return None



def pegar_debitos_fgts():
    mes_atual = datetime.now().strftime("%m-%Y")

    # Caminho para o diretório de downloads
    diretorio_download = os.path.join(os.getcwd(), f"debitos_fgts_{mes_atual}")

    # Cria o diretório se não existir
    if not os.path.exists(diretorio_download):
        os.makedirs(diretorio_download)

    # Configurações do Chrome para definir o diretório de download
    chrome_options = Options()
    chrome_options.add_argument(f'--download-default-directory={diretorio_download}')

    # Inicializa o WebDriver com as configurações
    driver = uc.Chrome(options=chrome_options)
    driver.get("https://fgtsdigital.sistema.gov.br/portal/login")
    driver.get("https://fgtsdigital.sistema.gov.br/portal/login")
    driver.maximize_window()

    sleep(2)

    entry_gov = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button is-primary entrar']"))
    )
    entry_gov.click()


    entry_certificate = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@id='login-certificate']"))
    )
    entry_certificate.click()
    sleep(5)

    try:
        hcaptcha_iframe = driver.find_element(By.XPATH, "//iframe[contains(@src, 'hcaptcha')]")
        if hcaptcha_iframe:
            print("hCaptcha detectado via Selenium.")
            driver.quit()
            sleep(10)
            pegar_debitos_fgts()
            return
    except Exception as e:
        # Se não encontrar, continua
        pass

    pyautogui.click(738,191, duration=2)
    sleep(5)

    pressionar_ate_encontrar(imagem_alvo, intervalo)

    sleep(3)

    try:
        aceitar_cookies = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button primary small']"))
        )
        aceitar_cookies.click()
    except Exception as e:
        print('Erro ao aceitar cookies', e)

    definir = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button is-primary']"))
    )
    definir.click()
    print('Clicado em definir')


    for linha in sheet_wb.iter_rows(min_row=2, max_row=500):
        cnpj = linha[2].value

        from selenium.common.exceptions import StaleElementReferenceException

        try:
            for _ in range(3):  # Tenta encontrar o botão até 3 vezes
                try:
                    trocar_perfil = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'botao-barra-perfil')]"))
                    )
                    trocar_perfil.click()
                    print('Clicado em "Trocar Perfil"')
                    break  # Sai do loop se conseguiu clicar
                except StaleElementReferenceException:
                    print("Elemento ficou obsoleto, tentando novamente...")
        except Exception as e:
            print(f"Erro ao clicar em 'Trocar Perfil': {e}")


        dropdown_perfil = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH,"//div[@role='combobox']"))
        )
        dropdown_perfil.click()

        procurador_option = driver.find_element(By.XPATH, "//span[@class='ng-option-label' and text()='Procurador']")

        # Clicando no item
        ActionChains(driver).move_to_element(procurador_option).click().perform()

        input_cnpj = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//input[@class='brx-input medium ng-untouched ng-pristine ng-invalid']"))
        )
        input_cnpj.click()
        input_cnpj.send_keys(cnpj)
        sleep(2)

        selecionar = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button is-primary']"))
        )
        selecionar.click()
        sleep(1)

        resultado_cnpj_nome = extrair_cnpj_nome_empresa(driver)
    
        if resultado_cnpj_nome:
            print(f"Resultado extraído: {resultado_cnpj_nome}")
        else:
            print("Não foi possível extrair o CNPJ e nome da empresa.")
        
        razao_social = resultado_cnpj_nome

        consultas_empregador = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//div[contains(@class, 'amplo cardListItem')]//span[contains(text(), 'Consultas do Empregador')]"))
        )
        consultas_empregador.click()
        sleep(2)

        consultar_competencia = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'amplo cardListItem')]//span[contains(text(), 'Consulta de Competências de Referência')]"))
        )

        # Clicar no elemento
        consultar_competencia.click()

        #elemento completo
        elemento = driver.find_element(By.XPATH, "//label[@for='completo']")
        actions = ActionChains(driver)
        actions.move_to_element(elemento).click().perform()
        sleep(1)

        #tirar o regular
        regular = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//label[@for='regular']"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(regular).click().perform()
        sleep(3)

        filtrar = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH,"//button[@class='br-button secondary']"))
        )
        filtrar.click()
        

        sleep(3)

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        sleep(2)  # Aguarde um pouco para garantir que o conteúdo foi carregado

        # Tentar localizar o elemento
        #elemento = WebDriverWait(driver, 10).until(
            #EC.element_to_be_clickable((By.CSS_SELECTOR, ".ng-arrow-wrapper"))
        #)
        elemento = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.XPATH, "(//div[@class='ng-input'])[1]"))
        )

        # Realizar a ação no elemento
        ActionChains(driver).move_to_element(elemento).click().perform()
        sleep(2)

        #checkbox_10 = WebDriverWait(driver, 10).until(
            #EC.element_to_be_clickable((By.XPATH, "(//div[@class='ng-value']//span[@class='ng-value-label' and text()='10'])[1]"))
        #)
        #checkbox_10.click()  # Clicando no valor "10"
        
        campo_texto = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//input[@aria-autocomplete='list']"))
        )

        # Clique no campo para abrir a lista
        campo_texto.click()

        # Digite o valor que você deseja selecionar (por exemplo, '100')
        campo_texto.send_keys("100")
        pyautogui.press('enter')
        sleep(3)

        dados = []

        # Encontra todas as linhas da tabela
        linhas = driver.find_elements(By.XPATH, "//datatable-body-row")

        for linha in linhas:
            try:
                # Extrai o mês de referência (supondo que seja a terceira célula)
                mes_ref = linha.find_element(By.XPATH, ".//datatable-body-cell[3]").text.strip()
                
                # Extrai o valor do débito (localizando pelo estilo de cor vermelha)
                valor_debito = linha.find_element(By.XPATH, ".//span[@style='color: #b30000; font-weight: 600;']").text.strip()

                # Remove caracteres indesejados no valor
                valor_debito = valor_debito.replace(".", "").replace(",", ".")

                # Adiciona os dados à lista
                dados.append([razao_social, mes_ref, float(valor_debito)])

            except Exception as e:
                print(f'Empresa {razao_social} sem débitos')
                print(f"Erro ao processar linha: {e}")

        # Criando um DataFrame com os novos dados
        df_novos_dados = pd.DataFrame(dados, columns=["Nome da Empresa", "Mês Ref.", "Valor Débitos"])

        # Caminho do arquivo Excel
        arquivo_excel = "debitos_fgts.xlsx"

        # Verificar se o arquivo já existe
        if os.path.exists(arquivo_excel):
            # Se o arquivo já existe, carregar os dados existentes
            df_existente = pd.read_excel(arquivo_excel)

            # Concatenar os dados novos com os dados existentes
            df_total = pd.concat([df_existente, df_novos_dados], ignore_index=True)

            # Salvar o DataFrame atualizado
            df_total.to_excel(arquivo_excel, index=False)
        else:
            # Se o arquivo não existe, criar um novo arquivo com os dados
            df_novos_dados.to_excel(arquivo_excel, index=False)

        print("Arquivo Excel atualizado com sucesso!")

        mes = datetime.now().strftime("%m-%Y")

        driver.get('https://fgtsdigital.sistema.gov.br/portal/servicos')


        sleep(2)
        
    driver.quit()
#pegar_debitos_fgts()

def criar_interface_fgts():
    # Limpa somente o conteúdo do info_frame (área dinâmica)
    for widget in info_frame.winfo_children():
        widget.destroy()

    # Cria um frame para a interface do FGTS Digital
    fgts_frame = ctk.CTkFrame(master=info_frame, width=400, height=500, corner_radius=20, fg_color="#2E2E2E")
    fgts_frame.place(relx=0.5, rely=0.5, anchor="center")

    # Título para a interface do FGTS Digital
    titulo_fgts = ctk.CTkLabel(
        master=fgts_frame,
        text="FGTS Digital",
        font=("Arial", 24, "bold"),
        text_color="#00A3FF"
    )
    titulo_fgts.pack(pady=20)

    # Instrução ou informações adicionais
    instrucao_fgts = ctk.CTkLabel(
        master=fgts_frame,
        text="Clique no botão abaixo para extrair débitos FGTS.",
        font=("Arial", 16),
        text_color="white"
    )
    instrucao_fgts.pack(pady=10)

    # Botão para chamar a função de extração dos débitos FGTS
    extrair_fgts_button = ctk.CTkButton(
        master=fgts_frame,
        text="Extrair Débitos FGTS",
        command=pegar_debitos_fgts,  # Função que você já definiu para extrair os débitos
        width=300,
        height=50,
        corner_radius=10,
        fg_color="#00A3FF",
        text_color="white",
        hover_color="#0088CC"
    )
    extrair_fgts_button.pack(pady=20)


#DIGILIZA
import customtkinter as ctk
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tkinter import messagebox
from PIL import Image
import pyautogui
from time import sleep
import pyperclip
import openpyxl
import os
import re
import pandas as pd
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from Levenshtein import distance

# Credenciais padrão
DEFAULT_EMAIL = "legal@contabilprimor.com.br"
DEFAULT_SENHA = "q7ne5k0la0VJ"


def extrair_nomes_empresas():
    pasta_debitos = "debitos"
    pasta_saida = "nomes_empresas"
    arquivo_saida = os.path.join(pasta_saida, "empresas.xlsx")

    # Garante que a pasta de saída existe
    os.makedirs(pasta_saida, exist_ok=True)

    dados = []

    # Iterar sobre os arquivos na pasta 'debitos'
    for arquivo in os.listdir(pasta_debitos):
        if arquivo.endswith(".pdf"):
            partes = arquivo.split("_")
            if len(partes) >= 2:
                cnpj = partes[0]
                nome_empresa = " ".join(partes[1:]).replace(".pdf", "").strip()
                dados.append([nome_empresa, cnpj])

    # Verifica se encontrou algum dado
    if not dados:
        print("Nenhum dado encontrado.")
        return

    # Criar um DataFrame e salvar como Excel
    df = pd.DataFrame(dados, columns=["Nome Empresa", "CNPJ"])
    df.to_excel(arquivo_saida, index=False)

    print(f"Arquivo salvo em: {arquivo_saida}")



# Função para iniciar o WebDriver com as credenciais
def iniciar_webdriver(email, senha):
    empresas_nao_enviadas = []
    total_empresas = contar_pdfs()  # Agora conta os arquivos PDF na pasta 'debitos'
    total_enviadas = 0
    #relatório de pendências fiscais
    #extrair_nomes_empresas()
    excel_msg = openpyxl.load_workbook(r"nomes_empresas\empresas.xlsx")
    sheet_excel_msg = excel_msg.active
    try:
        driver = webdriver.Chrome()
        driver.get("https://app.digiliza.com.br/login")
        driver.maximize_window()


        for linha in sheet_excel_msg.iter_rows(min_row=2, max_row=100):
            contador_vazios = 0
            if linha!=2:
                try:
                    driver.quit()  # Fecha antes de iniciar (se já estiver rodando)
                    # Inicia um novo WebDriver
                    driver = webdriver.Chrome()  
                    driver.get("https://app.digiliza.com.br/login")
                    driver.maximize_window()
                    input_email = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@id='email']"))
                    )
                    input_email.send_keys(email)

                    input_senha = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@id='password']"))
                    )
                    input_senha.send_keys(senha)

                    botao_login = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//button[@type='submit']"))
                    )
                    botao_login.click()
                except:
                    pass
            
    
            data_vcto_input_padrao = datetime.now().strftime("27/%m/%Y")
            print('Lopacarai',data_vcto_input_padrao)
            data_atual = datetime.now().strftime("%d/%m/%y")
            competencia_padrao = datetime.now().strftime("%m/%y")
            nome_empresa = linha[0].value
            cnpj = linha[1].value
            input_complemento = f'Análise de pendências da competência {data_atual}'

            if not nome_empresa or not cnpj:
                driver.quit()
                messagebox.showinfo('Sucesso!', 'Mensagens enviadas com sucesso!')
                break
            # Esperar até que o botão esteja presente
            incluir_tarefa_avulsa = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[@href='#modalTarefaAvulsa']"))
            )

            # Usar JavaScript para garantir o clique se necessário
            driver.execute_script("arguments[0].click();", incluir_tarefa_avulsa)

            sleep(3)

            select_model = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Selecione um Modelo']"))
            )
            select_model.click()
            sleep(2)

            texto = "relatório de pendências fiscais"
            pyperclip.copy(texto)
            pyautogui.hotkey("ctrl", "v")
            sleep(2)
            pyautogui.press('enter')

            cliente = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Selecione um Cliente']"))
            )
            cliente.click()
            cliente.send_keys(nome_empresa)
            sleep(2)
            pyautogui.press('enter')

            complemento = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@id='modalAddTaskComplemento']"))
            )
            complemento.click()
            complemento.send_keys(input_complemento)

            data_vcto_digiliza = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@type='date'][1]"))
            )
            data_vcto_digiliza.click()
            data_vcto_digiliza.clear()
            sleep(1)
            data_vcto_digiliza.send_keys(data_vcto_input_padrao)

            competencia = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='__/____']"))
            )
            competencia.click()
            competencia.clear()
            competencia.send_keys(competencia_padrao)
            
            '''
            #n precisa disso pq sempre vai ser primor contabil o responsavel
            responsavel = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//div[@class='v-select vs--single vs--searchable mh-sm scrollbar scrollbar-3']"))
            )
            responsavel.click()
            responsavel.clear()
            texto_responsavel = 'Prímor Contábil'
            pyperclip.copy(texto_responsavel)
            pyautogui.hotkey("ctrl", "v")
            print('dei ctrl v')
            sleep(15)
            pyautogui.press('enter')
            '''

            sleep(5)

            botao_salvar = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Salvar e ir à(s) Tarefa(s)')]"))
            )
            botao_salvar.click()
            sleep(3)

            # Supondo que 'driver' já esteja inicializado
            try:
                # Espera até 2 segundos para encontrar o título do erro
                WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.XPATH, "//h2[@id='swal2-title' and contains(text(), 'Erro ao criar tarefa(s)')]"))
                )
                
                # Se encontrar, espera o botão "OK" e clica nele
                btn_ok = WebDriverWait(driver, 2).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'swal2-confirm')]"))
                )
                btn_ok.click()
                print("Botão 'OK' clicado com sucesso.")

                empresas_nao_enviadas.append(f"{nome_empresa} ({cnpj})")
                continue
            except Exception as e:  
                if "WinError 6" in str(e) or "chrome not reachable" in str(e):
                    print("ChromeDriver fechou inesperadamente. Reiniciando...")
                    driver.quit()
                    driver = webdriver.Chrome()  # Inicia um novo WebDriver
                    print("Mensagem de erro não encontrada dentro do tempo limite.")


            #Acessando o elemento com nome da empresa incrementado c f''
            

            # Captura todas as empresas da tabela
            nomes_encontrados = [e.text for e in driver.find_elements(By.XPATH, "//td")]
            print('Nomes encontrados\n')
            # Encontra o nome mais próximo
            nome_correto = min(nomes_encontrados, key=lambda x: distance(x.lower(), nome_empresa.lower()))
            print('Nome correto', nome_correto)

            # Clica no nome encontrado
            elemento_empresa = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, f"//td[contains(text(), '{nome_correto}')]"))
            )
            elemento_empresa.click()

            sleep(3)

            botao_documentos_svg = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-transparent rounded-0 text-primary flex-fill']"))
            )
            botao_documentos_svg.click()
            print('Acessei a aba de documentos')
            sleep(3)

            # Diretório onde estão os PDFs
            pasta_debitos = os.path.join(os.getcwd(), 'debitos')

            # Construir padrão do nome do arquivo
            padrao_arquivo = f"{cnpj}_*.pdf"

            # Procurar o arquivo correto na pasta
            arquivos_encontrados = [f for f in os.listdir(pasta_debitos) if f.startswith(f"{cnpj}_") and f.endswith(".pdf")]

            if arquivos_encontrados:
                caminho_arquivo = os.path.join(pasta_debitos, arquivos_encontrados[0])  # Pega o primeiro encontrado


                # Localizar o campo de upload oculto e enviar o arquivo
                campo_upload = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@type='file']"))
                )
                campo_upload.send_keys(caminho_arquivo)

                print(f"Arquivo {caminho_arquivo} anexado com sucesso!")

                try:
                    aceitar_pop_up = WebDriverWait(driver,3).until(
                        EC.element_to_be_clickable((By.XPATH,"//button[@class='swal2-confirm px-3 py-2 swal2-styled swal2-default-outline']"))
                    )
                    aceitar_pop_up.click()

                except:
                    print('n tinha pop up')
                sleep(2)
            else:
                print(f"Nenhum arquivo encontrado para {nome_empresa} ({cnpj})")

            # Esperar o upload ser processado (caso tenha carregamento)
            sleep(5) 

            enviar_msg = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-sm btn-transparent px-1 py-0'] [1]"))
            )
            enviar_msg.click()

            click_edit_mensagem = WebDriverWait(driver,5).until(
                EC.element_to_be_clickable((By.XPATH,"//div[@class='ql-editor ql-blank']"))
            )
            click_edit_mensagem.click()

            # Abrir o arquivo mensagens.xlsx
            excel_mensagens = openpyxl.load_workbook("mensagens.xlsx")
            sheet_mensagens = excel_mensagens.active

            # Procurar o CNPJ na coluna 0 e pegar a mensagem da coluna 1
            mensagem_personalizada = None
            for linha in sheet_mensagens.iter_rows(min_row=2, max_row=10):
                cnpj_planilha = str(linha[0].value)
                print(f'Cnpj {cnpj_planilha}')
                mensagem = str(linha[1].value)
                #print(f'Mensagem {mensagem}')
                if str(cnpj).strip() == str(cnpj_planilha).strip():  # Comparar sem espaços extras
                    mensagem_personalizada = mensagem
                    break

            # Se encontrou a mensagem, colar no campo de edição
            if mensagem_personalizada:
                click_edit_mensagem = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, "//div[@class='ql-editor ql-blank']"))
                )
                click_edit_mensagem.click()

                # Copiar e colar usando pyperclip + pyautogui
                click_edit_mensagem.send_keys(mensagem_personalizada)
                print("Colei")
                sleep(2)

                botoes = driver.find_elements(By.XPATH, "//button[text()='Salvar']")
                for botao in botoes:
                    if botao.is_displayed():
                        botao.click()
                        break

                print('Cliquei em salvar')
                sleep(5)
                
                voltar = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-transparent rounded-0 text-secondary flex-fill']"))
                )
                voltar.click()

                iniciar_etapa = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-sm btn-transparent btn-hover-semi-transparent text-white']"))
                )
                iniciar_etapa.click()
                sleep(5)

                concluir_etapa = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-sm btn-transparent btn-hover-semi-transparent text-white big-chungus']"))
                )
                concluir_etapa.click()
                sleep(3)

                botao_ok = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='swal2-confirm px-3 py-2 swal2-styled']"))
                )
                botao_ok.click()

                total_enviadas += 1

                '''
                concluir_e_enviar = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[@class='btn btn-success' and contains(text(), 'Concluir e enviar')]"))
                )
                concluir_e_enviar.click()

                botao_ok = WebDriverWait(driver,5).until(
                    EC.element_to_be_clickable((By.XPATH,"//button[@class='swal2-confirm px-3 py-2 swal2-styled']"))
                )
                botao_ok.click()
                '''
                botoes_fechar = driver.find_elements(By.XPATH, "//button[@data-bs-dismiss='modal']")
                for botao in botoes_fechar:
                    
                    if botao.is_displayed():  # Verifica se o botão está visível na tela
                        botao.click()
                        break  # Para após clicar no primeiro botão visível

                

                sleep(5)

                print(f"Mensagem enviada para {nome_empresa}: {mensagem_personalizada}")
            else:
                empresas_nao_enviadas.append(f"{nome_empresa} ({cnpj})")
                '''
                botoes_fechar = driver.find_elements(By.XPATH, "//button[@data-bs-dismiss='offcanvas']")
                for botao in botoes_fechar:
                    print('Nao achei a mensagem!')
                    if botao.is_displayed():  # Verifica se o botão está visível na tela
                        botao.click()
                        break  # Para após clicar no primeiro botão visível
                '''
                print(f"⚠ Nenhuma mensagem personalizada encontrada para {nome_empresa} ({cnpj})")

        if empresas_nao_enviadas:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Nome da Empresa", "CNPJ"])
            for empresa in empresas_nao_enviadas:
                nome, cnpj = empresa.split(" (")
                cnpj = cnpj[:-1]  # Removendo o parêntese final
                ws.append([nome, cnpj])
    
            wb.save("empresas_nao_enviadas.xlsx")
            print("Arquivo 'empresas_nao_enviadas.xlsx' salvo com sucesso.")

        # Enviar email com os resultados
        enviar_email(empresas_nao_enviadas, total_empresas, total_enviadas)
        messagebox.showinfo("Sucesso", "Login realizado com sucesso!")
        sleep(50)

        if driver.service.is_connectable():
            driver.quit()
    except Exception as e:
        print("Erro", f"Falha no login: {e}")

def contar_pdfs():
    pasta = "debitos"
    return len([f for f in os.listdir(pasta) if f.endswith(".pdf")])

def enviar_email(empresas_nao_enviadas, total_empresas, total_enviadas):
    remetente_email = "luizhill.dev@gmail.com"  # Altere para seu email
    senha_email = "nqlf fgch thrs kpht"  # Use app password se for Gmail
    destinatarios = ["fiscal@contabilprimor.com.br", "pessoal@contabilprimor.com.br"]

    saudacao = "Bom dia" if datetime.now().hour < 12 else "Boa tarde"
    lista_empresas = "\n".join(empresas_nao_enviadas)

    mensagem = f"""{saudacao},  
    Informo que as seguintes empresas não foram enviadas:  
    {lista_empresas}  

    Havia um total de {total_empresas} empresas (arquivos PDF na pasta 'debitos'), foram enviadas {total_enviadas}.
    """

    msg = MIMEMultipart()
    msg["From"] = remetente_email
    msg["To"] = ", ".join(destinatarios)
    msg["Subject"] = "Empresas não enviadas"

    msg.attach(MIMEText(mensagem, "plain"))

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)  # Alterar se não for Gmail
        server.starttls()
        server.login(remetente_email, senha_email)
        server.sendmail(remetente_email, destinatarios, msg.as_string())
        server.quit()
        print("Email enviado com sucesso!")
    except Exception as e:
        print(f"Erro ao enviar email: {e}")


eye_open = ctk.CTkImage(Image.open(r"imgs\eye_open.png"), size=(24, 24))
eye_closed = ctk.CTkImage(Image.open(r"imgs\eye_closed.png"), size=(24, 24))

def criar_interface():
    global eye_open, eye_closed  # As imagens devem ser criadas globalmente

    def fazer_login():
        email = email_entry.get()
        senha = senha_entry.get()
        iniciar_webdriver(email, senha)

    def toggle_password():
        """Alterna entre exibir ou ocultar a senha"""
        if senha_entry.cget("show") == "*":
            senha_entry.configure(show="")
            toggle_button.configure(image=eye_open)
        else:
            senha_entry.configure(show="*")
            toggle_button.configure(image=eye_closed)

    # Limpa apenas o conteúdo do info_frame
    for widget in info_frame.winfo_children():
        widget.destroy()

    # Cria a interface de login dentro do info_frame
    login_frame = ctk.CTkFrame(master=info_frame, width=400, height=500, corner_radius=20, fg_color="#2E2E2E")
    login_frame.place(relx=0.5, rely=0.5, anchor="center")

    titulo = ctk.CTkLabel(master=login_frame, text="Login no Digiliza", font=("Arial", 24, "bold"), text_color="#00A3FF")
    titulo.pack(pady=20)

    email_label = ctk.CTkLabel(master=login_frame, text="E-mail:", text_color="white")
    email_label.pack()
    email_entry = ctk.CTkEntry(master=login_frame, width=300, height=40, corner_radius=10)
    email_entry.insert(0, DEFAULT_EMAIL)
    email_entry.pack(pady=5)

    senha_label = ctk.CTkLabel(master=login_frame, text="Senha:", text_color="white")
    senha_label.pack()

    # Frame para o campo de senha e botão de exibição
    senha_frame = ctk.CTkFrame(master=login_frame, fg_color="transparent")
    senha_frame.pack()

    senha_entry = ctk.CTkEntry(master=senha_frame, width=260, height=40, corner_radius=10, show="*")
    senha_entry.insert(0, DEFAULT_SENHA)
    senha_entry.pack(side="left", pady=5)

    toggle_button = ctk.CTkButton(
        master=senha_frame,
        width=40,
        height=40,
        text="",
        image=eye_closed,  # Imagem padrão (senha oculta)
        fg_color="transparent",
        hover_color="#444",
        command=toggle_password
    )
    toggle_button.image = eye_closed  # Mantém a referência da imagem
    toggle_button.pack(side="right", padx=5)

    login_button = ctk.CTkButton(
        master=login_frame,
        text="Login",
        command=fazer_login,
        width=300,
        height=50,
        corner_radius=10,
        fg_color="#00A3FF",
        text_color="white",
        hover_color="#0088CC"
    )
    login_button.pack(pady=20)


# Caminho para a pasta "resultados"
diretorio_resultados = os.path.join(os.getcwd(), 'resultados')
# Caminhos para as pastas e arquivos
diretorio_codigos = os.path.join(os.getcwd(), 'resultados_codigos')
arquivo_tabelas = os.path.join(os.getcwd(), 'TABELASCDIGOSDERECEITA.xlsx')
tabela_depto_pessoal = pd.read_excel(arquivo_tabelas, sheet_name='Depto Pessoal')
tabela_fiscal = pd.read_excel(arquivo_tabelas, sheet_name='Fiscal')



def salvar_mensagem(df_existente, nome_empresa, nova_mensagem, caminho_saida):
    # Lista de empresas já existentes no arquivo
    nomes_existentes = df_existente['Empresa'].tolist()

    # Mostrar os nomes das empresas existentes no DataFrame
    print("Empresas existentes no arquivo:", nomes_existentes)

    # Encontrar o nome mais parecido
    nome_mais_proximo, score = process.extractOne(nome_empresa, nomes_existentes) if nomes_existentes else (None, 0)

    # Mostrar o nome mais próximo e o score
    print(f"Procurando pelo nome: {nome_empresa}")
    print(f"Nome mais próximo encontrado: {nome_mais_proximo}, Score: {score}")

    # Se encontrou uma correspondência confiável, usa o nome existente
    if nome_mais_proximo and score >= 93:
        nome_empresa = nome_mais_proximo
        print(f"Usando nome mais próximo: {nome_empresa}")

    # Se a empresa já existir no arquivo, concatena a mensagem
    if nome_empresa in df_existente['Empresa'].values:
        print(f"Empresa '{nome_empresa}' encontrada no arquivo, concatenando mensagem...")
        df_existente.loc[df_existente['Empresa'] == nome_empresa, 'Mensagem'] += f"\n{nova_mensagem}"
    else:
        print(f"Empresa '{nome_empresa}' não encontrada no arquivo, criando nova linha...")
        nova_linha = pd.DataFrame({"Empresa": [nome_empresa], "Mensagem": [nova_mensagem]})
        df_existente = pd.concat([df_existente, nova_linha], ignore_index=True)

    return df_existente



diretorio_processos_sief = os.path.join(os.getcwd(), 'processos sief')

def criar_msg_fgts():
    # Carregar os arquivos
    fgts_df = pd.read_excel("debitos_fgts.xlsx")
    mensagens_df = pd.read_excel("mensagens.xlsx")

    # Criar um dicionário para agrupar os débitos por empresa
    fgts_dict = {}
    for _, row in fgts_df.iterrows():
        nome_completo = row["Nome da Empresa"]
        cnpj, nome_empresa = nome_completo.split("_", 1)
        mes_ref = row["Mês Ref."]
        valor = row["Valor Débitos"]
        
        if cnpj not in fgts_dict:
            fgts_dict[cnpj] = {"nome": nome_empresa, "debitos": {}}
        
        if mes_ref not in fgts_dict[cnpj]["debitos"]:
            fgts_dict[cnpj]["debitos"][mes_ref] = 0
        
        fgts_dict[cnpj]["debitos"][mes_ref] += valor

    # Criar ou atualizar as mensagens
    for cnpj, data in fgts_dict.items():
        nome_empresa = data["nome"]
        debitos_texto = ", ".join([f"{mes}: R$ {valor:.2f}" for mes, valor in data["debitos"].items()])
        
        if cnpj in mensagens_df["Empresa"].astype(str).values:
            print('tinha o cnpj', cnpj)
            mensagem_fgts = f"{nome_empresa}, você também possui débitos de FGTS: " + ", ".join(
                [f"{mes} no valor de R$ {valor:.2f}" for mes, valor in data['debitos'].items()]
            ) + "."
            mensagens_df.loc[mensagens_df["Empresa"].astype(str) == cnpj, "Mensagem"] += f" {mensagem_fgts}"
        else:
            mensagem = f"{nome_empresa}, segue resumo dos seus débitos de FGTS: {debitos_texto}."
            mensagens_df = pd.concat([mensagens_df, pd.DataFrame({"Empresa": [cnpj], "Mensagem": [mensagem]})], ignore_index=True)

    # Salvar o arquivo atualizado
    mensagens_df.to_excel("mensagens.xlsx", index=False)

    print("Mensagens de FGTS geradas e salvas com sucesso!")

def criar_msgs_processos_sief(caminho_saida, diretorio_processos_sief):
    from datetime import datetime
    import os
    import pandas as pd

    data_atual = datetime.now().strftime("%d/%m/%y")
    
    # Verifica se já existe um arquivo com mensagens e carrega os dados
    if os.path.exists(caminho_saida):
        df_existente = pd.read_excel(caminho_saida)
    else:
        df_existente = pd.DataFrame(columns=["Empresa", "Mensagem"])

    # Percorre todos os arquivos Excel na pasta
    for arquivo in os.listdir(diretorio_processos_sief):
        if arquivo.endswith('.xlsx') or arquivo.endswith('.xls'):
            # Exemplo do nome do arquivo: "23098061000139_C R V ESTERO E CIA LTDA.xlsx"
            # Extrai o CNPJ que está antes do primeiro '_'
            cnpj = arquivo.split('_')[0]
            print(f"🔍 CNPJ extraído do nome do arquivo: {cnpj}")

            # Monta o caminho completo do arquivo
            caminho_arquivo = os.path.join(diretorio_processos_sief, arquivo)
            
            # Lê o arquivo Excel
            df = pd.read_excel(caminho_arquivo)
            
            # Verifica se a coluna necessária existe
            if 'Processos SIEF' in df.columns:
                # Lista de processos (removendo valores vazios)
                processos = df["Processos SIEF"].dropna().astype(str).tolist()

                match = re.match(r'^\d+_(.*)\.xlsx$', arquivo)
                if match:
                    nome_empresa_sem_cnpj = match.group(1)
                else:
                    nome_empresa_sem_cnpj = "Nome não encontrado"
                
                # Gera a mensagem personalizada usando somente o CNPJ como identificador
                mensagem = f"\n\nA empresa {nome_empresa_sem_cnpj} possui os seguintes débitos referentes a Processos SIEF:\n\n"
                mensagem += ', '.join(processos)
                
                # Salva ou concatena a mensagem no DataFrame existente, usando o CNPJ como chave
                df_existente = salvar_mensagem(df_existente, cnpj, mensagem.strip(), caminho_saida)
                
                print(f"✅ Mensagem gerada para {cnpj}:\n{mensagem}\n")
            else:
                print(f"⚠️ O arquivo {arquivo} não possui a coluna 'Processos SIEF' esperada.")

    # Salva as mensagens geradas no arquivo Excel
    df_existente.to_excel(caminho_saida, index=False)
    print("✅ Mensagens salvas com sucesso!")

def criar_msgs(caminho_saida):
    data_atual = datetime.now().strftime("%d/%m/%y")
    
    # Percorre todos os arquivos Excel na pasta
    for arquivo in os.listdir(diretorio_resultados):

        if os.path.exists(caminho_saida):
            df_existente = pd.read_excel(caminho_saida)
        else:
            df_existente = pd.DataFrame(columns=["Empresa", "Mensagem"])

        if arquivo.endswith('.xlsx') or arquivo.endswith('.xls'):  # Verifica se é um arquivo Excel
            caminho_arquivo = os.path.join(diretorio_resultados, arquivo)
            
            # Lê o arquivo Excel
            df = pd.read_excel(caminho_arquivo)
            
            # Garante que as colunas necessárias estão no DataFrame
            if {'EMPRESA', 'DÍVIDA ATIVA', 'NUMERO DO PROCESSO', 'SITUAÇÃO'}.issubset(df.columns):
                
                # Tenta extrair o CNPJ limpo (14 dígitos) da coluna "EMPRESA"
                cnpj = re.search(r'(\d{14})', str(df['EMPRESA'].iloc[0]))  # Supondo que o CNPJ esteja na primeira linha
                if cnpj:
                    cnpj = cnpj.group(1)  # Extrai o CNPJ limpo
                    
                    # Remover o CNPJ do nome da empresa para utilizá-lo na mensagem
                    nome_empresa_sem_cnpj = df['EMPRESA'].iloc[0].replace(cnpj + "_", "")  # Remove o CNPJ do início do nome
                    
                    print(f"🔍 Buscando pelo CNPJ: {cnpj}")
                    
                    # Agrupa os processos pela mesma situação
                    situacoes = df.groupby('SITUAÇÃO')['NUMERO DO PROCESSO'].apply(list).to_dict()
                    
                    # Gera a mensagem personalizada para a empresa (usando o nome sem o CNPJ)
                    mensagem = f"\n\nA empresa possui os seguintes débitos na Procuradoria-Geral da Fazenda Nacional: \n"
                    for situacao, processos in situacoes.items():
                        processos_formatados = ', '.join(processos)  # Junta os números dos processos
                        mensagem += f"{situacao}'.\n"

                    df_existente = salvar_mensagem(df_existente, cnpj, mensagem.strip(), caminho_saida)
                    
                    print(f"Mensagem para {nome_empresa_sem_cnpj}:\n{mensagem}\n")
                else:
                    print(f"⚠️ CNPJ não encontrado para a empresa '{df['EMPRESA'].iloc[0]}'.")
            else:
                print(f"O arquivo {arquivo} não possui as colunas esperadas.")
            
        df_existente.to_excel(caminho_saida, index=False)
        print("Mensagens salvas com sucesso!")



def criar_msgs_codigos(diretorio_codigos, tabela_depto_pessoal, tabela_fiscal, caminho_saida):
    data_atual = datetime.now().strftime("%d/%m/%y")
    if os.path.exists(caminho_saida):
        df_existente = pd.read_excel(caminho_saida)
    else:
        df_existente = pd.DataFrame(columns=["Empresa", "Mensagem"])

    for arquivo in os.listdir(diretorio_codigos):
        if arquivo.endswith('.xlsx') or arquivo.endswith('.xls'):
            caminho_arquivo = os.path.join(diretorio_codigos, arquivo)
            df = pd.read_excel(caminho_arquivo)

            if {'Empresa', 'Código Fiscal', 'PA - Exercício', 'Saldo Devedor Consignado'}.issubset(df.columns):
                # Tenta extrair o CNPJ limpo (14 dígitos) da coluna "Empresa"
                cnpj = re.search(r'(\d{14})', str(df['Empresa'].iloc[0]))  # Supondo que o CNPJ esteja na primeira linha
                if cnpj:
                    cnpj = cnpj.group(1)  # Extrai o CNPJ limpo
                    
                    # Remover o CNPJ do nome da empresa para utilizá-lo na mensagem
                    nome_empresa_sem_cnpj = df['Empresa'].iloc[0].replace(cnpj + "_", "")  # Remove o CNPJ do início do nome
                    
                    print(f"🔍 Buscando pelo CNPJ: {cnpj}")
                    
                    # Função para ajustar o formato do PA - Exercício
                    def formatar_pa_exercicio(pa_exercicio):
                        try:
                            if len(str(pa_exercicio).split('/')) == 3:  # Caso DDD/MM/YYYY
                                return '/'.join(str(pa_exercicio).split('/')[1:])
                            return str(pa_exercicio)
                        except Exception as e:
                            print(f"Erro ao formatar PA - Exercício: {pa_exercicio}, erro: {e}")
                            return None

                    df['PA - Exercício'] = df['PA - Exercício'].apply(formatar_pa_exercicio)

                    # Agrupando os dados pelo PA - Exercício
                    meses_agrupados = df.groupby('PA - Exercício')

                    mensagem = f"Olá {nome_empresa_sem_cnpj},\n"
                    mensagem += "Identificamos que sua empresa possui algumas pendências em aberto junto à Receita Federal.\n"
                    mensagem += "Essas pendências podem gerar multas, juros e complicações mais sérias se não forem regularizadas em tempo hábil.\n\n"
                    mensagem += "Segue o resumo dos seus débitos:\n\n"
                    mensagem = textwrap.dedent(mensagem)




                    for pa_exercicio, grupo in meses_agrupados:
                        mensagem += f"**Referente a {pa_exercicio}:**\n"
                        debitos_por_tipo = {}

                        for _, row in grupo.iterrows():
                            codigo_fiscal_completo = str(row['Código Fiscal']).strip()
                            saldo_devedor = str(row['Saldo Devedor Consignado']).replace(',', '.')

                            try:
                                saldo_devedor = float(saldo_devedor)
                            except ValueError:
                                saldo_devedor = 0.0  # Caso o valor não seja numérico, considera como zero

                            if saldo_devedor <= 0:
                                continue  # Ignora débitos zerados

                            match = re.match(r'(\d+)[-/](\d+)', codigo_fiscal_completo)
                            if match:
                                codigo_fiscal_formatado_original = f"{match.group(1)}-{match.group(2)}"
                                codigo_fiscal_com_variacao = f"{match.group(1)}/{match.group(2)}"
                            else:
                                codigo_fiscal_formatado_original = codigo_fiscal_completo
                                codigo_fiscal_com_variacao = codigo_fiscal_completo

                            # Verifica em qual tabela o código está presente
                            if (codigo_fiscal_formatado_original in tabela_depto_pessoal['Código de receita'].astype(str).values or
                                codigo_fiscal_com_variacao in tabela_depto_pessoal['Código de receita'].astype(str).values):
                                tipo_debito = "Departamento Pessoal"
                            elif (codigo_fiscal_formatado_original in tabela_fiscal['Código de receita'].astype(str).values or
                                  codigo_fiscal_com_variacao in tabela_fiscal['Código de receita'].astype(str).values):
                                tipo_debito = "Fiscal"
                            else:
                                descricao = re.sub(r'^\d+[-/]\d+\s-\s', '', codigo_fiscal_completo)
                                tipo_debito = f"outros ({descricao})"

                            # Soma os valores por tipo de débito
                            if tipo_debito in debitos_por_tipo:
                                debitos_por_tipo[tipo_debito] += saldo_devedor
                            else:
                                debitos_por_tipo[tipo_debito] = saldo_devedor

                        # Adiciona os valores somados à mensagem
                        for tipo, valor in debitos_por_tipo.items():
                            mensagem += f"  - {tipo}: R$ {valor:.2f}\n"
                        
                        mensagem += "\n"  # Separação entre meses

                    df_existente = salvar_mensagem(df_existente, cnpj, mensagem.strip(), caminho_saida)
                    print(f"Mensagem gerada para {nome_empresa_sem_cnpj}:\n{mensagem}\n")
                else:
                    print(f"⚠️ CNPJ não encontrado para a empresa '{df['Empresa'].iloc[0]}'.")
            else:
                print(f"O arquivo {arquivo} não possui as colunas esperadas.")

    df_existente.to_excel(caminho_saida, index=False)
    print("Mensagens salvas com sucesso!")


# Chamada da função
import pandas as pd

def criar_msg_fgts():
    # Carregar os arquivos
    fgts_df = pd.read_excel("debitos_fgts.xlsx")
    mensagens_df = pd.read_excel("mensagens.xlsx")

    # Criar um dicionário para agrupar os débitos por empresa
    fgts_dict = {}
    for _, row in fgts_df.iterrows():
        nome_completo = row["Nome da Empresa"]
        cnpj, nome_empresa = nome_completo.split("_", 1)
        mes_ref = row["Mês Ref."]
        valor = row["Valor Débitos"]
        
        if cnpj not in fgts_dict:
            fgts_dict[cnpj] = {"nome": nome_empresa, "debitos": {}}
        
        if mes_ref not in fgts_dict[cnpj]["debitos"]:
            fgts_dict[cnpj]["debitos"][mes_ref] = 0
        
        fgts_dict[cnpj]["debitos"][mes_ref] += valor

    # Criar ou atualizar as mensagens
    for cnpj, data in fgts_dict.items():
        nome_empresa = data["nome"]
        debitos_texto = ", ".join([f"{mes}: R$ {valor:.2f}" for mes, valor in data["debitos"].items()])
        
        if cnpj in mensagens_df["Empresa"].astype(str).values:
            print('tinha o cnpj', cnpj)
            mensagem_fgts = f"{nome_empresa}, você também possui débitos de FGTS: " + ", ".join(
                [f"{mes} no valor de R$ {valor:.2f}" for mes, valor in data['debitos'].items()]
            ) + "."
            mensagens_df.loc[mensagens_df["Empresa"].astype(str) == cnpj, "Mensagem"] += f" {mensagem_fgts}"
        else:
            mensagem = f"{nome_empresa}, segue resumo dos seus débitos de FGTS: {debitos_texto}."
            mensagens_df = pd.concat([mensagens_df, pd.DataFrame({"Empresa": [cnpj], "Mensagem": [mensagem]})], ignore_index=True)

    # Salvar o arquivo atualizado
    mensagens_df.to_excel("mensagens.xlsx", index=False)

    print("Mensagens de FGTS geradas e salvas com sucesso!")


def criar_msg_final():
    # Carregar o arquivo de mensagens
    mensagens_df = pd.read_excel("mensagens.xlsx")

    # Definir a mensagem final
    data_atual = datetime.now().strftime("%d/%m/%y")
    mensagem_final = (
        f"\nOs valores informados são válidos na data de envio deste e-mail ({data_atual}) e podem sofrer alterações.\n"
        "Caso tenha interesse em regularizar essas pendências, entre em contato com o nosso time "
        "para mais detalhes e orientações sobre os próximos passos.\n"
        "Ficamos à disposição para qualquer dúvida ou informação adicional!\n\n"
        "Atenciosamente,\n"
        "Prímor Contábil\n"
        "(44) 98462-9927 / atendimento@contabilprimor.com.br"
    )

    # Garantir que a mensagem final seja a última coisa adicionada a cada linha
    mensagens_df["Mensagem"] = mensagens_df["Mensagem"].astype(str) + mensagem_final

    # Salvar as mensagens atualizadas
    mensagens_df.to_excel("mensagens.xlsx", index=False)

    print("Mensagem final adicionada com sucesso!")





def extrair_nome_empresa_e_cnpj(nome_arquivo):
    """
    Extrai o nome da empresa e o CNPJ do nome do arquivo PDF.
    O nome do arquivo segue o formato 'situacao_fiscal--CNPJ-Nome_Arquivo.pdf'.
    """
    # Expressão regular para capturar o CNPJ
    cnpj = re.search(r"situacao_fiscal--(\d{14})-", nome_arquivo)
    if cnpj:
        cnpj = cnpj.group(1)  # Extrai o CNPJ

    # Remove o prefixo 'situacao_fiscal--CNPJ-' e qualquer código no final
    nome_limpo = re.sub(r"situacao_fiscal--\d{14}-", "", nome_arquivo)
    nome_limpo = re.sub(r"_[0-9]+\.pdf$", "", nome_limpo)  # Remove código final (se existir)
    
    return nome_limpo.strip(), cnpj

def renomear_pdfs_com_cnpj(pasta):
    """
    Itera sobre todos os PDFs da pasta e renomeia, usando o CNPJ como ID único.
    """
    for arquivo in os.listdir(pasta):
        if arquivo.endswith(".pdf"):
            caminho_antigo = os.path.join(pasta, arquivo)
            nome_empresa, cnpj = extrair_nome_empresa_e_cnpj(arquivo)
            
            # Criar o novo nome do arquivo com CNPJ
            if cnpj:
                novo_nome = f"{cnpj}_{nome_empresa}.pdf"
                caminho_novo = os.path.join(pasta, novo_nome)
                
                # Renomeia o arquivo
                os.rename(caminho_antigo, caminho_novo)
                print(f"Renomeado: {arquivo} → {novo_nome}")
            else:
                print(f"Não foi possível extrair o CNPJ de: {arquivo}")

# Defina a pasta onde estão os PDFs
pasta_pdfs = "debitos"

def consultar_pdf_da_empresa(nome_empresa, numeros_procurados):
    # Caminho da pasta onde os PDFs foram baixados/descompactados
    pasta_debitos = os.path.join(os.getcwd(), 'debitos')

    # Listar todos os arquivos PDF na pasta
    arquivos_pdf = [f for f in os.listdir(pasta_debitos) if f.endswith('.pdf')]

    # Tentar extrair o CNPJ limpo (sem pontos, barras ou traços) do nome da empresa
    cnpj = re.search(r'(\d{14})', nome_empresa)
    
    if cnpj:
        cnpj = cnpj.group(1)  # Extrai o CNPJ limpo
        print(f"🔍 Buscando pelo CNPJ: {cnpj}")
        
        # Buscar diretamente o PDF com base no CNPJ limpo
        nome_pdf_proximo = None
        for arquivo in arquivos_pdf:
            if cnpj in arquivo:  # Verifica se o CNPJ está no nome do arquivo PDF
                nome_pdf_proximo = arquivo
                break

        if nome_pdf_proximo:
            caminho_pdf = os.path.join(pasta_debitos, nome_pdf_proximo)
            print(f"🔍 PDF encontrado: {nome_pdf_proximo}")

            # Aqui você pode usar uma biblioteca para abrir o PDF, por exemplo, pdfplumber
            abrir_pdf(caminho_pdf, numeros_procurados, nome_empresa)
        else:
            print(f"⚠️ Nenhum PDF encontrado para o CNPJ '{cnpj}'.")
    else:
        print(f"⚠️ CNPJ não encontrado no nome da empresa '{nome_empresa}'.")

#busca os numeros de dividas ativas aqui


def abrir_pdf(caminho_pdf, numeros_procurados, nome_empresa, pasta_destino="resultados"):
    """
    Abre o PDF, procura por números específicos, combina texto de todas as páginas 
    e captura a situação associada ao número.
    """
    resultados = []  # Lista para armazenar os dados processados

    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            # Combinar o texto de todas as páginas do PDF
            texto_completo = ""
            for page in pdf.pages:
                texto_completo += page.extract_text() + "\n"

            # Procurar os números no texto combinado
            for numero in numeros_procurados:
                if str(numero) in texto_completo:
                    print(f"⚠️ Número encontrado no PDF: {numero}")

                    # Verificar qualquer situação associada ao número
                    padrao = rf"{numero}.*?Situação:\s+([^\n]+)"
                    match = re.search(padrao, texto_completo, re.DOTALL)
                    
                    if match:
                        situacao = match.group(1).strip()
                        print(f"✅ Situação do número {numero}: {situacao}")
                        
                        # Adiciona os dados na lista
                        resultados.append({
                            "EMPRESA": nome_empresa,
                            "DÍVIDA ATIVA": "SIM",
                            "NUMERO DO PROCESSO": numero,
                            "SITUAÇÃO": situacao
                        })
                    else:
                        print(f"⚠️ Situação do número {numero} não encontrada.")
        
        # Criar a pasta de destino se não existir
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        
        # Salvar resultados no Excel
        if resultados:
            caminho_excel = os.path.join(pasta_destino, f"{nome_empresa}_resultados.xlsx")
            df = pd.DataFrame(resultados)
            df.to_excel(caminho_excel, index=False)
            print(f"✅ Resultados salvos em: {caminho_excel}")
        else:
            print(f"⚠️ Nenhum dado encontrado para a empresa: {nome_empresa}")

    except Exception as e:
        print(f"Erro ao abrir ou processar o arquivo PDF {caminho_pdf}: {e}")



def processar_excel_e_abrir_pdf():
    """
    Processa os arquivos Excel na pasta 'dividas ativas', extrai os números das dívidas e
    busca por esses números nos PDFs relacionados.
    """
    pasta_destino = "dividas ativas"
    
    for excel_file in os.listdir(pasta_destino):
        if excel_file.endswith('.xlsx'):
            # Caminho completo do arquivo Excel
            caminho_excel = os.path.join(pasta_destino, excel_file)

            # Ler os números do Excel
            df = pd.read_excel(caminho_excel)
            numeros_procurados = df["Inscrição da Dívida"].astype(str).tolist()
            
            # Extrair o nome da empresa a partir do nome do arquivo Excel
            nome_empresa = os.path.splitext(excel_file)[0].replace(" LTDA", "")
            print(f"Procurando números no PDF para a empresa: {nome_empresa}")

            # Localizar o PDF correspondente
            consultar_pdf_da_empresa(nome_empresa, numeros_procurados)

#CONSULTA PDF DOS CODIGOS FISCAIS
def consultar_pdf_da_empresa_codigos(nome_empresa, numeros_procurados):
    """
    Procura o PDF correspondente ao nome da empresa e chama a função para extrair os valores.
    """
    # Caminho da pasta onde os PDFs estão
    pasta_debitos = os.path.join(os.getcwd(), 'debitos')

    # Listar todos os arquivos PDF na pasta
    arquivos_pdf = [f for f in os.listdir(pasta_debitos) if f.endswith('.pdf')]

    # Tentar extrair o CNPJ limpo (sem pontos, barras ou traços) do nome da empresa
    cnpj = re.search(r'(\d{14})', nome_empresa)
    
    if cnpj:
        cnpj = cnpj.group(1)  # Extrai o CNPJ limpo
        print(f"🔍 Buscando pelo CNPJ: {cnpj}")
        
        # Buscar diretamente o PDF com base no CNPJ limpo
        nome_pdf_proximo = None
        for arquivo in arquivos_pdf:
            if cnpj in arquivo:  # Verifica se o CNPJ está no nome do arquivo PDF
                nome_pdf_proximo = arquivo
                break

        if nome_pdf_proximo:
            caminho_pdf = os.path.join(pasta_debitos, nome_pdf_proximo)
            print(f"🔍 PDF encontrado: {nome_pdf_proximo}")

            # Chamar a função para processar o PDF
            abrir_pdf_codigos(caminho_pdf, numeros_procurados, nome_empresa)
        else:
            print(f"⚠️ Nenhum PDF encontrado para o CNPJ '{cnpj}'.")
    else:
        print(f"⚠️ CNPJ não encontrado no nome da empresa '{nome_empresa}'.")

def abrir_pdf_codigos(caminho_pdf, numeros_procurados, nome_empresa):
    """
    Abre o PDF, busca os códigos fiscais e PA - EXERC. e extrai o saldo devedor consignado.
    """
    resultados = []  # Lista para armazenar os resultados

    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            # Combinar o texto de todas as páginas do PDF
            texto_completo = ""
            for page in pdf.pages:
                texto_completo += page.extract_text() + "\n"

            # Iterar pelos códigos fiscais e PA-EXERC. fornecidos
            for codigo, pa_exerc in numeros_procurados:
                # Regex para encontrar a linha correspondente no PDF
                padrao = rf"{codigo}\s+{pa_exerc}.*?([\d.,]+)\s*DEVEDOR$"
                match = re.search(padrao, texto_completo, re.MULTILINE)

                if match:
                    saldo_devedor = match.group(1).strip()
                    print(f"✅ Código: {codigo}, PA-EXERC.: {pa_exerc}, Saldo: {saldo_devedor}")
                    
                    # Adicionar o resultado
                    resultados.append({
                        "Empresa": nome_empresa,
                        "Código Fiscal": codigo,
                        "PA - Exercício": pa_exerc,
                        "Saldo Devedor Consignado": saldo_devedor
                    })
                else:
                    print(f"⚠️ Nenhuma linha encontrada para Código: {codigo}, PA-EXERC.: {pa_exerc}")

        # Salvar os resultados em um Excel
        if resultados:
            salvar_resultados_excel(resultados, nome_empresa)
        else:
            print(f"⚠️ Nenhum dado extraído do PDF para a empresa: {nome_empresa}")

    except Exception as e:
        print(f"Erro ao abrir ou processar o PDF {caminho_pdf}: {e}")



def salvar_resultados_excel(resultados, nome_empresa):
    """
    Salva os resultados extraídos em um arquivo Excel.
    """
    # Criar um DataFrame com os resultados
    df = pd.DataFrame(resultados)

    # Caminho para salvar os resultados
    pasta_destino = os.path.join(os.getcwd(), 'resultados_codigos')
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)

    caminho_excel = os.path.join(pasta_destino, f"{nome_empresa}_codigos_fiscais.xlsx")
    df.to_excel(caminho_excel, index=False)
    print(f"✅ Resultados salvos em: {caminho_excel}")


def processar_empresas_codigos():
    """
    Itera sobre os arquivos Excel na pasta 'codigos fiscais' e processa os PDFs correspondentes.
    """
    pasta_codigos = os.path.join(os.getcwd(), 'codigos fiscais')

    for arquivo_excel in os.listdir(pasta_codigos):
        if arquivo_excel.endswith('.xlsx'):
            caminho_excel = os.path.join(pasta_codigos, arquivo_excel)
            nome_empresa = os.path.splitext(arquivo_excel)[0]

            # Ler o Excel para obter os códigos fiscais e PA-EXERC.
            df = pd.read_excel(caminho_excel)

            numeros_procurados = list(zip(df["Codigos Fiscais"], df["PA - EXERC."]))

            print(f"🔍 Processando empresa: {nome_empresa}")
            consultar_pdf_da_empresa_codigos(nome_empresa, numeros_procurados)


#salva os debitos de divida ativa
def salvar_numeros_em_excel(lista_numeros, nome_arquivo, pasta_destino):
    # Salvar a lista de números em um arquivo Excel
    df = pd.DataFrame(lista_numeros, columns=["Inscrição da Dívida"])
    caminho_arquivo = os.path.join(pasta_destino, f"{nome_arquivo}.xlsx")
    df.to_excel(caminho_arquivo, index=False)
    print(f"Arquivo salvo em {caminho_arquivo}")

import pandas as pd
import os

def salvar_processos_em_excel(lista_processos, nome_arquivo, pasta_destino, nome_empresa):
    # Criar o DataFrame
    df = pd.DataFrame(lista_processos, columns=["Processos SIEF"])

    # Adicionar a coluna "Nome Empresa" na primeira posição
    df.insert(0, "Nome Empresa", nome_empresa)  

    # Criar caminho do arquivo
    caminho_arquivo = os.path.join(pasta_destino, f"{nome_arquivo}.xlsx")
    
    # Salvar no Excel
    df.to_excel(caminho_arquivo, index=False)
    print(f"Arquivo salvo em {caminho_arquivo}")

# Chamando a função corretamente


#codigos fiscais salvando

def salvar_codigos_em_excel(lista_numeros, lista_pa_exercicio, nome_arquivo, pasta_codigos):
    # Verifique se as listas têm o mesmo comprimento
    if len(lista_numeros) != len(lista_pa_exercicio):
        raise ValueError("As listas de números e de PA - EXERC. devem ter o mesmo comprimento.")

    # Verificar se o diretório existe, se não, criar o diretório
    if not os.path.exists(pasta_codigos):
        os.makedirs(pasta_codigos)
    
    # Salvar a lista de números e PA em um DataFrame
    df = pd.DataFrame({
        "Codigos Fiscais": lista_numeros,  # Lista de números
        "PA - EXERC.": lista_pa_exercicio  # Lista de PA - Exercicio
    })
    
    # Definir o caminho do arquivo Excel
    caminho_arquivo = os.path.join(pasta_codigos, f"{nome_arquivo}.xlsx")
    
    # Salvar o DataFrame no arquivo Excel
    df.to_excel(caminho_arquivo, index=False)
    
    print(f"Arquivo salvo em {caminho_arquivo}")

def descompactar_arquivo_zip(download_folder, driver):

    zip_file = None

    while not zip_file:
        # Verifica se há um arquivo ZIP na pasta
        for file in os.listdir(download_folder):
            if file.endswith('.zip'):
                zip_file = os.path.join(download_folder, file)
                break  # Sai do loop assim que encontrar o primeiro ZIP
                
        if not zip_file:
            pyautogui.press('f5')
            sleep(2)
            baixar_nuvem = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(@id, '__BV_toggle_') and contains(@class, 'nav-link dropdown-toggle')]"))
            )
            baixar_nuvem.click()
            print("Nenhum arquivo ZIP encontrado. Tentando novamente em 5 segundos...")
            pyautogui.click(667, 300, duration=1)  # Clica no botão
            time.sleep(10)  # Espera 10 segundos antes de tentar novamente

    # Se encontrou um ZIP, descompacta
    try:
        with zipfile.ZipFile(zip_file, 'r') as zip_ref:
            zip_ref.extractall(download_folder)
        print(f"Arquivo ZIP {zip_file} descompactado.")

        # Excluir o arquivo ZIP
        os.remove(zip_file)
        print(f"Arquivo ZIP {zip_file} excluído.")
    except Exception as e:
        print(f"Erro ao descompactar ou excluir o arquivo ZIP: {e}")

# Caminho da pasta onde os PDFs foram descompactados
pasta_debitos = os.path.join(os.getcwd(), 'debitos')

# Função para carregar a tabela de códigos fiscais
def carregar_codigos_fiscais(caminho_arquivo_excel):
    df_depto = pd.read_excel(caminho_arquivo_excel, sheet_name="Depto Pessoal")
    df_fiscal = pd.read_excel(caminho_arquivo_excel, sheet_name="Fiscal")

    # Concatenar os dois dataframes
    df = pd.concat([df_depto, df_fiscal])

    # Criar um dicionário {codigo: descricao}
    codigos_fiscais = dict(zip(df.iloc[:, 0].astype(str), df.iloc[:, 1]))
    
    return codigos_fiscais

import shutil

def limpar_pastas():
    pastas = ['debitos', 'resultados', 'resultados_codigos', 'codigos fiscais', 'dividas ativas', 'processos sief']
    
    for pasta in pastas:
        if os.path.exists(pasta):
            for arquivo in os.listdir(pasta):
                caminho_arquivo = os.path.join(pasta, arquivo)
                if os.path.isfile(caminho_arquivo) or os.path.islink(caminho_arquivo):
                    os.unlink(caminho_arquivo)  # Remove arquivos e links simbólicos
                elif os.path.isdir(caminho_arquivo):
                    shutil.rmtree(caminho_arquivo)  # Remove subpastas
            print(f"Conteúdo da pasta '{pasta}' foi excluído.")
        else:
            print(f"A pasta '{pasta}' não existe.")



def login():
    limpar_pastas()
    download_folder = os.path.join(os.getcwd(), 'debitos')

    # Configuração para o Chrome salvar os PDFs diretamente na pasta especificada
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": download_folder,  # Caminho para a pasta onde o PDF será salvo
        "download.prompt_for_download": False,  # Desativa o prompt para confirmar download
        "plugins.always_open_pdf_externally": True  # Abre os PDFs diretamente sem pedir para abrir
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=options)
    driver.get('https://app.monitorcontabil.com.br/login')
    driver.maximize_window()


    try:
        botao_verde = WebDriverWait(driver,3).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@class='swal2-confirm swal-btn-green swal2-styled']"))
        )
        botao_verde.click()
    except:
        print('N tinha botao verde')

    sleep(2)

    logar_email = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//input[@id='email']"))
    )
    logar_email.send_keys('taina@contabilprimor.com.br')

    logar_senha = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//input[@id='senhaInput']"))
    )
    logar_senha.send_keys('Primor1214')

    logar = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@type='submit']"))
    )
    logar.click()

    sleep(3)
    pyautogui.press('esc')

    driver.get("https://app.monitorcontabil.com.br/situacao-fiscal/visualizar?busca=")

    sleep(2)

    
    filtro_irregulares = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, "//select[@name='filtroSelecao']"))
    )

    # Abre a lista de opções do select
    filtro_irregulares.click()

    # Seleciona a opção que contém "Irregulares"
    select = Select(filtro_irregulares)

    # Aqui vamos selecionar a opção "Irregulares"
    for option in select.options:
        if "Irregulares" in option.text:
            option.click()
            sleep(1)
            pyautogui.press('esc')
            break
        
    sleep(3)

    baixar_relatorios = WebDriverWait(driver, 15).until(
    EC.element_to_be_clickable((By.XPATH, "//button[.//span[text()='Baixar situações']]"))
    )
    baixar_relatorios.click()


    baixar_popup = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH,"//button[@class='btn btn-sm btn-success']"))
    )
    baixar_popup.click()

    sleep(8)
    pyautogui.press('F5')
    sleep(2)

    baixar_nuvem = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//a[contains(@id, '__BV_toggle_') and contains(@class, 'nav-link dropdown-toggle')]"))
    )
    baixar_nuvem.click()


    WebDriverWait(driver, 5).until(
        EC.invisibility_of_element_located((By.XPATH, "//small[@class='notification-text' and text()='Em Execução']"))
    )

    # Agora que o "Em Execução" não está mais visível, podemos clicar no botão de download
    #baixar_definitivamente = WebDriverWait(driver, 10).until(
        #EC.element_to_be_clickable((By.XPATH, "//div[@class='media mr-1']//div[@class='col-2']//svg[contains(@class, 'feather-download')]"))
    #)
    #baixar_definitivamente.click()


    pyautogui.click(667,300, duration = 1)

    sleep(5)
    descompactar_arquivo_zip(download_folder, driver)
    sleep(2)
    renomear_pdfs_com_cnpj(pasta_pdfs)
    pyautogui.press('Esc')
    sleep(2)

    filtro_a_z = WebDriverWait(driver,5).until(
        EC.element_to_be_clickable((By.XPATH, "//th[contains(@class, 'vgt-left-align') and contains(@class, 'sortable')]/span[text()='Razão social']"))
    )
    filtro_a_z.click()
    sleep(2)


    pasta_cnpjs = os.path.join(os.getcwd(), 'cnpj_empresas')
    if not os.path.exists(pasta_cnpjs):
        os.makedirs(pasta_cnpjs)

    # Nome do arquivo Excel
    arquivo_excel_cnpjs = 'empresas_cnpj.xlsx'
    caminho_excel_cnpjs = os.path.join(pasta_cnpjs, arquivo_excel_cnpjs)

    # Lista para armazenar os dados das empresas
    dados_cnpjs = []

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "tr.clickable")))

    # Encontrar todas as linhas da tabela
    while True:
        linhas = driver.find_elements(By.CSS_SELECTOR, "tr.clickable")

        for i, linha in enumerate(linhas):  # Define o início da iteração
            linhas = driver.find_elements(By.CSS_SELECTOR, "tr.clickable")
            print("Numero de linhas:",len(linhas))
            print(f'Processando linha {i + 1} de {len(linhas)}')
            try:
                print(f'Processando linha {linhas.index(linha) + 1} de {len(linhas)}')
                # Extrair o CNPJ
                cnpj = linha.find_element(By.CSS_SELECTOR, "td.vgt-left-align.col-tamanho-cnpj span span span").text
                cnpj = re.sub(r'[^0-9]', '', cnpj)
                # Extrair o Nome da Empresa
                nome_empresa = linha.find_element(By.CSS_SELECTOR, "td.vgt-left-align span span span").text

                print(f"Empresa: {nome_empresa} | CNPJ: {cnpj}")

                # Encontrar e clicar na lupa correspondente
                lupa = linha.find_element(By.XPATH, ".//button[@class='btn btn btn-none rounded-pill m-0 icone-acao p-0 btn-none btn-none'][2]")
                
                # Usar JavaScript para garantir o clique, caso necessário
                driver.execute_script("arguments[0].click();", lupa)

                # Aguarde o carregamento da nova informação, se houver
                WebDriverWait(driver, 5).until(EC.staleness_of(lupa))  # Espera a lupa desaparecer/mudar
            except Exception as e:
                print(f"Erro ao processar uma linha: {e}")

                #tentando clicar em débitos(sief)
                try:    
                    debitos_sief = WebDriverWait(driver,5).until(
                        EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'list-group-item') and contains(@class, 'collapsed')]//span[text()='Débito (Sief)']"))
                    )
                    debitos_sief.click()
                    print('CLICADO EM DEBITOS SIEF')
                    numeros = driver.find_elements(By.XPATH, "//tr//td[@aria-colindex='1']//div[@class='ml-50']")
                    lista_numeros = [numero.text for numero in numeros]

                    print("Códigos fiscais:", lista_numeros, "\n")

                    pa_exercicio = driver.find_elements(By.XPATH, "//tr//td[@aria-colindex='2']//div[@class='ml-50']")
                    lista_pa_exercicio = [pa.text for pa in pa_exercicio]
                    print("Pa - exercício = ", lista_pa_exercicio, "\n")

                    nome_arquivo = f"{cnpj}_{nome_empresa}".replace("/", "_").replace(".", "_")  # Substituindo caracteres não permitidos
                    pasta_debitos = os.path.join(os.getcwd(), 'debitos')
                    pasta_codigos = "codigos fiscais"
                    salvar_codigos_em_excel(lista_numeros, lista_pa_exercicio, nome_arquivo, pasta_codigos)
                    processar_empresas_codigos()

                except Exception as e:
                    print('n cliquei em debitos sie F')
                    print(f'Erro{e}')

                #tentando clicar em dívidas ativas
                try: 
                    divida_ativa = WebDriverWait(driver,5).until(
                        EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'list-group-item') and contains(@class, 'collapsed')]//span[text()='Dividas ativas']"))
                    )
                    divida_ativa.click()

                    # Extrair todos os números de todos os <div class="ml-50"> dentro da coluna específica
                    numeros = driver.find_elements(By.XPATH, "//tr//td[@aria-colindex='1']//div[@class='ml-50']")
                    lista_numeros = [numero.text for numero in numeros]

        
                    # Gerar nome do arquivo com CNPJ e nome da empresa
                    nome_arquivo = f"{cnpj}_{nome_empresa}".replace("/", "_").replace(".", "_")  # Substituindo caracteres não permitidos

                    pasta_destino = "dividas ativas"
                    if not os.path.exists(pasta_destino):
                        os.makedirs(pasta_destino)
                    
                    salvar_numeros_em_excel(lista_numeros, nome_arquivo, pasta_destino)
                    sleep(1)
                    processar_excel_e_abrir_pdf()
                    #PEGAR os numeros da dívida, passar pra um excel, ai a partir disso ir no pdf e extrair
                    #Todos que estiverem com ''Pendencia - inscrição'', situação ''Ativa em cobrança'' ou ''Ativa a ser cobrada'', precisa colocar pois são pendências em divida ativa que não foram negociadas ainda
                    #Poderia colocar na mensagem algo como: Pendência em Inscrição em dívida ativa na Procuradoria-Geral da Fazenda Nacional:- colocar os números das inscrições e data que foi inscrito (obs: quando estiver parcelamento rescindido não aparecerá data da inscrição)
                    #Os que estiverem em ''Inscrição com Exigibilidade Suspensa'' E ''Parcelamento com Exigibilidade Suspensa'' não precisa informar nada, pois as vidas já estão negociadas e parceladas

                except Exception as e:
                    print(f"Erro ao clicar em 'Dividas Ativas' ou extrair os números: {e}")

                #tentando clicar em processo fiscal sief
                try: 
                    processo_sief = WebDriverWait(driver,2).until(
                        EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'list-group-item') and contains(@class, 'collapsed')]//span[text()='Processo Fiscal (Sief)']"))
                    )
                    processo_sief.click()

                    # Extrair todos os números de todos os <div class="ml-50"> dentro da coluna específica
                    processos = driver.find_elements(By.XPATH, "//tr//td[@aria-colindex='1']//div[@class='ml-50']")
                    lista_processos = [processo.text for processo in processos]

                    # Gerar nome do arquivo com CNPJ e nome da empresa
                    nome_arquivo = f"{cnpj}_{nome_empresa}".replace("/", "_").replace(".", "_")  # Substituindo caracteres não permitidos

                    pasta_destino_sief = "processos sief"
                    if not os.path.exists(pasta_destino_sief):
                        os.makedirs(pasta_destino_sief)
                    
                    salvar_processos_em_excel(lista_processos, nome_arquivo, pasta_destino_sief, nome_empresa)
                    sleep(1)
                    #PEGAR os numeros da dívida, passar pra um excel, ai a partir disso ir no pdf e extrair
                    #Todos que estiverem com ''Pendencia - inscrição'', situação ''Ativa em cobrança'' ou ''Ativa a ser cobrada'', precisa colocar pois são pendências em divida ativa que não foram negociadas ainda
                    #Poderia colocar na mensagem algo como: Pendência em Inscrição em dívida ativa na Procuradoria-Geral da Fazenda Nacional:- colocar os números das inscrições e data que foi inscrito (obs: quando estiver parcelamento rescindido não aparecerá data da inscrição)
                    #Os que estiverem em ''Inscrição com Exigibilidade Suspensa'' E ''Parcelamento com Exigibilidade Suspensa'' não precisa informar nada, pois as vidas já estão negociadas e parceladas

                except Exception as e:
                    print(f"Erro ao clicar em 'Dividas Ativas' ou extrair os números: {e}")
                
                

                pyautogui.press('esc')
                sleep(2)

                

                if i + 1 == len(linhas):
                    try:
                        sleep(1)
                        pula_pagina = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Go to next page']"))
                        )
                        pula_pagina.click()
                        print("Avançando para a próxima página...")
                        linhas = driver.find_elements(By.CSS_SELECTOR, "tr.clickable")
                        print("Numero de linhas:",len(linhas))
                    except Exception as e:
                        print(f"Erro ao clicar no botão de avançar página: {e}")
                        driver.quit()
        
                        
            pasta_debitos = os.path.join(os.getcwd(), 'debitos')


def atualizar_informacoes(frame):
    """
    Simula a atualização de informações dinâmicas no frame central.
    """
    for widget in frame.winfo_children():
        widget.destroy()  # Limpa o conteúdo anterior

    info_label = ctk.CTkLabel(
        frame,
        text="Atualizando informações... 🚀",
        font=("Arial", 16),
        anchor="center",
    )
    info_label.pack(pady=10)

    # Exemplo de atualização futura
    info_atualizada = ctk.CTkLabel(
        frame,
        text="Extração de débitos em andamento...",
        font=("Arial", 14),
        anchor="center",
    )
    info_atualizada.pack(pady=5)

import customtkinter as ctk
from tkinter import messagebox
from datetime import datetime
import schedule
import time
import threading

def criar_msgs_geral():
    criar_msgs_codigos(diretorio_codigos, tabela_depto_pessoal, tabela_fiscal, caminho_saida = 'mensagens.xlsx')
    criar_msgs(caminho_saida="mensagens.xlsx")
    criar_msgs_processos_sief(caminho_saida="mensagens.xlsx", diretorio_processos_sief = diretorio_processos_sief)
    criar_msg_fgts()
    criar_msg_final()

# Função para configurar o horário
def agendar_robo():
    horario = horario_entry.get()  # Pegando o valor do campo de entrada
    try:
        # Validar o formato de hora
        horario_formatado = datetime.strptime(horario, "%H:%M")
        
        # Agendar a execução do login para o horário escolhido
        horario_str = horario_formatado.strftime("%H:%M")
        schedule.every().day.at(horario_str).do(login)  # Executa a função login todos os dias nesse horário

        messagebox.showinfo("Horário agendado", f"O robô será executado às {horario_formatado.strftime('%H:%M')}")
        
        # Iniciar o agendamento em uma thread separada para não bloquear a interface
        threading.Thread(target=run_schedule).start()

    except ValueError:
        messagebox.showerror("Erro", "Formato de hora inválido! Use o formato HH:MM.")

# Função para rodar o schedule em uma thread separada
def run_schedule():
    while True:
        schedule.run_pending()  # Verifica se alguma tarefa agendada precisa ser executada
        time.sleep(1)  # Espera 1 segundo antes de verificar novamente


# Configuração inicial da interface
ctk.set_appearance_mode("dark")  # "System", "Light" ou "Dark"
ctk.set_default_color_theme("blue")  # "blue", "green" ou "dark-blue"

# Janela principal
app = ctk.CTk()
app.geometry("900x600")
app.title("Sistema de Extração de Débitos - Primor")

# Menu lateral
menu_frame = ctk.CTkFrame(app, width=250, corner_radius=0)
menu_frame.pack(side="left", fill="y")

menu_label = ctk.CTkLabel(
    menu_frame,
    text="Menu",
    font=("Arial", 20, "bold"),
    anchor="w",
)
menu_label.pack(pady=(20, 10), padx=10, anchor="w")

# Botão no menu
extrair_button = ctk.CTkButton(
    menu_frame,
    text="Extrair débitos",
    command=login,
    font=("Arial", 16, "bold"),
)
extrair_button.pack(pady=20, padx=10)

botton_msgs = ctk.CTkButton(
    menu_frame,
    text="Criar mensagens",
    command=criar_msgs_geral,
    font=("Arial", 16, "bold"),
)
botton_msgs.pack(pady=20, padx=10)

digitaliza = ctk.CTkButton(
    menu_frame,
    text="Enviar mensagens",
    command=criar_interface,  # Essa função atualizará somente o conteúdo dinâmico
    font=("Arial", 16, "bold"),
)
digitaliza.pack(pady=20, padx=10)

fgts_digital_button = ctk.CTkButton(
    menu_frame,
    text="FGTS Digital",
    command=criar_interface_fgts,  # Atualiza o conteúdo dinâmico para FGTS Digital
    font=("Arial", 16, "bold")
)
fgts_digital_button.pack(pady=20, padx=10)


# Área principal
main_frame = ctk.CTkFrame(app, corner_radius=10)
main_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)

# Cabeçalho (FIXO)
main_label = ctk.CTkLabel(
    main_frame,
    text="Bem-vindo ao PrimorFiscal Messenger",
    font=("Arial", 24, "bold"),
    anchor="center",
)
main_label.pack(pady=20)

# (Opcional) Outros elementos fixos podem ser adicionados aqui, como imagens e parcerias...
# Por exemplo, o frame de imagens e o label de parceria:
try:
    from PIL import Image, ImageTk
    primor_image = ImageTk.PhotoImage(Image.open("imgs/primor.png").resize((200, 120)))
    luiz_image = ImageTk.PhotoImage(Image.open("imgs/luiz.png").resize((200, 120)))

    images_frame = ctk.CTkFrame(main_frame)
    images_frame.pack(pady=20)

    primor_label = ctk.CTkLabel(images_frame, image=primor_image, text="")
    primor_label.grid(row=0, column=0, padx=10)

    luiz_label = ctk.CTkLabel(images_frame, image=luiz_image, text="")
    luiz_label.grid(row=0, column=1, padx=10)

    partnership_label = ctk.CTkLabel(
        main_frame,
        text="Uma parceria entre Primor e Luiz Fernando Hillebrande",
        font=("Arial", 16, "italic"),
        anchor="center",
    )
    partnership_label.pack(pady=(10, 20))
except Exception as e:
    print(f"Erro ao carregar imagens: {e}")

# Área dinâmica para atualizações (essa área será atualizada pela função criar_interface)
info_frame = ctk.CTkFrame(main_frame, height=200)
info_frame.pack(fill="both", expand=True, padx=10, pady=(10, 20))

# Seleção de horário (fora do main_frame, se preferir)
horario_label = ctk.CTkLabel(
    app, text="Escolha o horário para o robô rodar (formato HH:MM):", font=("Arial", 16)
)
horario_label.pack(pady=10)

horario_entry = ctk.CTkEntry(app, width=150, font=("Arial", 14))
horario_entry.insert(0, "09:00")  # Horário padrão
horario_entry.pack(pady=10)

# Botão para agendar
agendar_button = ctk.CTkButton(
    main_frame,
    text="Agendar Robô",
    command=agendar_robo,
    font=("Arial", 16, "bold"),
)
agendar_button.pack(pady=10)

# Footer (centralizado)
footer_frame = ctk.CTkFrame(app, height=50, corner_radius=0)
footer_frame.pack(side="bottom", fill="x")

footer_label = ctk.CTkLabel(
    footer_frame,
    text="Luiz Fernando Hillebrande",
    font=("Arial", 14),
    anchor="center",
)
footer_label.pack(pady=10)

# Iniciar a interface
app.mainloop()
